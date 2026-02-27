"""
BestLife Hub - FastAPI Backend
Handles: Auth verification, TherapyNotes upload/processing, analytics, user management, invoices.
"""
import os
import io
import json
import logging
import calendar
from collections import OrderedDict
from datetime import datetime, timedelta, date
from typing import Optional, List

import httpx
import openpyxl
from fastapi import FastAPI, UploadFile, File, HTTPException, Depends, Request, Body
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel

# ────────────────────────────────────────────────────────────────────
# Config
# ────────────────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://jvtwvrqityxzcnsbrilk.supabase.co")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
SUPABASE_ANON_KEY = os.environ.get("SUPABASE_ANON_KEY", "")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "") or os.environ.get("CLAUDE_API_KEY", "")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("bestlife")

app = FastAPI(title="BestLife Hub API")

# Startup: load secrets from Supabase if not in env vars
@app.on_event("startup")
async def startup_event():
    global ANTHROPIC_API_KEY
    logger.info("BestLife Hub API starting up...")
    logger.info(f"Supabase URL: {SUPABASE_URL}")
    logger.info(f"Service key configured: {'Yes' if SUPABASE_SERVICE_KEY else 'No'}")

    # If Anthropic key not in env vars, try loading from Supabase app_settings
    if not ANTHROPIC_API_KEY and SUPABASE_SERVICE_KEY:
        try:
            async with httpx.AsyncClient(timeout=10.0) as client:
                resp = await client.get(
                    f"{SUPABASE_URL}/rest/v1/app_settings",
                    params={"key": "eq.ANTHROPIC_API_KEY", "select": "value"},
                    headers={
                        "apikey": SUPABASE_SERVICE_KEY,
                        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
                    },
                )
                if resp.status_code == 200:
                    rows = resp.json()
                    if rows and rows[0].get("value"):
                        ANTHROPIC_API_KEY = rows[0]["value"]
                        logger.info(f"Loaded ANTHROPIC_API_KEY from Supabase app_settings (len={len(ANTHROPIC_API_KEY)})")
                    else:
                        logger.warning("app_settings table exists but no ANTHROPIC_API_KEY row found")
                else:
                    logger.warning(f"Could not read app_settings: {resp.status_code}")
        except Exception as e:
            logger.warning(f"Failed to load API key from Supabase: {e}")

    if ANTHROPIC_API_KEY:
        logger.info(f"Anthropic key ready: Yes (len={len(ANTHROPIC_API_KEY)}, prefix={ANTHROPIC_API_KEY[:8]}...)")
    else:
        logger.warning("Anthropic key: NOT configured (Betty AI will be unavailable)")


# ────────────────────────────────────────────────────────────────────
# Supabase helpers
# ────────────────────────────────────────────────────────────────────
def sb_headers(service=False):
    """Headers for Supabase REST API calls."""
    key = SUPABASE_SERVICE_KEY if service else SUPABASE_ANON_KEY
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }


async def sb_request(method, path, data=None, params=None, service=True):
    """Make a request to Supabase REST API."""
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = sb_headers(service=service)
    async with httpx.AsyncClient(timeout=30.0) as client:
        resp = await client.request(method, url, json=data, params=params, headers=headers)
        if resp.status_code >= 400:
            logger.error(f"Supabase error {resp.status_code}: {resp.text}")
            raise HTTPException(status_code=resp.status_code, detail=resp.text)
        if resp.status_code == 204:
            return None
        return resp.json()


async def verify_token(request: Request) -> dict:
    """Verify Supabase JWT and return user profile."""
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing authorization token")

    token = auth.split(" ", 1)[1]

    # Verify with Supabase Auth
    async with httpx.AsyncClient(timeout=10.0) as client:
        resp = await client.get(
            f"{SUPABASE_URL}/auth/v1/user",
            headers={
                "apikey": SUPABASE_ANON_KEY or SUPABASE_SERVICE_KEY,
                "Authorization": f"Bearer {token}",
            },
        )
        if resp.status_code != 200:
            raise HTTPException(status_code=401, detail="Invalid token")
        auth_user = resp.json()

    # Get user profile from users table
    users = await sb_request("GET", "users", params={"auth_id": f"eq.{auth_user['id']}", "select": "*"})
    if not users:
        raise HTTPException(status_code=403, detail="User profile not found")

    return users[0]


async def require_admin(request: Request) -> dict:
    """Verify user is admin."""
    profile = await verify_token(request)
    if profile.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return profile


# ────────────────────────────────────────────────────────────────────
# Models
# ────────────────────────────────────────────────────────────────────
class InviteUserRequest(BaseModel):
    email: str
    first_name: str
    last_name: str
    role: str
    employment_status: Optional[str] = "full_time"
    phone_number: Optional[str] = None
    sms_enabled: bool = True
    supervision_required: bool = False
    clinical_supervisor_id: Optional[str] = None


class AIChatRequest(BaseModel):
    prompt: str
    context: Optional[str] = None
    system_hint: Optional[str] = None
    max_tokens: int = 1024


# ────────────────────────────────────────────────────────────────────
# API Routes
# ────────────────────────────────────────────────────────────────────

@app.get("/api/health")
async def health():
    return {"status": "ok", "version": "1.0.0"}


# ── User Management ──────────────────────────────────────────────

@app.post("/api/admin/invite-user")
async def invite_user(req: InviteUserRequest, admin=Depends(require_admin)):
    """Create a new user via Supabase Auth invite and add to users table."""
    # Step 1: Invite via Supabase Auth — this creates the auth user AND sends the email
    async with httpx.AsyncClient(timeout=15.0) as client:
        invite_resp = await client.post(
            f"{SUPABASE_URL}/auth/v1/invite",
            headers={
                "apikey": SUPABASE_SERVICE_KEY,
                "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "email": req.email,
                "data": {
                    "first_name": req.first_name,
                    "last_name": req.last_name,
                },
            },
        )

        if invite_resp.status_code >= 400:
            err_body = invite_resp.json() if invite_resp.headers.get("content-type", "").startswith("application/json") else {}
            error_detail = err_body.get("msg") or err_body.get("error_description") or err_body.get("message") or invite_resp.text
            raise HTTPException(status_code=400, detail=f"Auth error: {error_detail}")

        auth_user = invite_resp.json()
        auth_id = auth_user.get("id", "")

    # Step 2: Insert into users table
    user_data = {
        "auth_id": auth_id,
        "email": req.email,
        "first_name": req.first_name,
        "last_name": req.last_name,
        "role": req.role,
        "employment_status": req.employment_status or "full_time",
        "is_active": True,
    }
    if req.phone_number:
        user_data["phone_number"] = req.phone_number
    if req.sms_enabled is not None:
        user_data["sms_enabled"] = req.sms_enabled
    if req.supervision_required is not None:
        user_data["supervision_required"] = req.supervision_required
    if req.clinical_supervisor_id:
        user_data["clinical_supervisor_id"] = req.clinical_supervisor_id

    new_user = await sb_request("POST", "users", data=user_data)
    user_id = new_user[0]["id"] if isinstance(new_user, list) else new_user.get("id")

    return {"status": "invited", "email": req.email, "user_id": user_id}


# ── TherapyNotes Upload & Processing ────────────────────────────

@app.post("/api/upload/therapynotes")
async def upload_therapynotes(file: UploadFile = File(...), admin=Depends(require_admin)):
    """Process a TherapyNotes billing export and store transactions in Supabase."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active

    # Parse headers
    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {h: i for i, h in enumerate(headers)}

    required_cols = ["Record Type"]
    for col in required_cols:
        if col not in header_map:
            raise HTTPException(status_code=400, detail=f"Missing required column: {col}")

    # Parse rows
    transactions = []
    therapist_names = set()
    date_min = None
    date_max = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        def cell(name):
            idx = header_map.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        record_type = str(cell("Record Type") or "").strip()
        patient_name = str(cell("Patient Name") or cell("Patient") or "").strip()
        provider_name = str(cell("Provider Name") or cell("Provider") or "").strip()
        service_date = cell("Service Date") or cell("Date")
        amount = cell("Amount") or cell("Payment Amount") or 0
        payer = str(cell("Payer") or cell("Payer Name") or "").strip()
        code = str(cell("Code") or cell("Procedure Code") or "").strip()
        description = str(cell("Description") or "").strip()

        # Parse date
        date_str = None
        if isinstance(service_date, datetime):
            date_str = service_date.strftime("%Y-%m-%d")
        elif service_date:
            try:
                for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
                    try:
                        date_str = datetime.strptime(str(service_date).strip(), fmt).strftime("%Y-%m-%d")
                        break
                    except ValueError:
                        continue
            except Exception:
                pass

        # Parse amount
        try:
            amount_val = float(str(amount).replace("$", "").replace(",", "").replace("(", "-").replace(")", ""))
        except (ValueError, TypeError):
            amount_val = 0.0

        if provider_name:
            therapist_names.add(provider_name)

        if date_str:
            d = datetime.strptime(date_str, "%Y-%m-%d")
            if date_min is None or d < date_min:
                date_min = d
            if date_max is None or d > date_max:
                date_max = d

        transactions.append({
            "record_type": record_type,
            "patient_name": patient_name,
            "provider_name": provider_name,
            "service_date": date_str,
            "amount": amount_val,
            "payer": payer,
            "code": code,
            "description": description,
        })

    wb.close()

    if not transactions:
        raise HTTPException(status_code=400, detail="No transactions found in file")

    # Batch insert into Supabase (chunks of 500)
    chunk_size = 500
    for i in range(0, len(transactions), chunk_size):
        chunk = transactions[i:i + chunk_size]
        await sb_request("POST", "transactions", data=chunk)

    # Ensure therapists exist in users table (upsert check)
    for name in therapist_names:
        parts = name.split(None, 1)
        first = parts[0] if parts else name
        last = parts[1] if len(parts) > 1 else ""

        existing = await sb_request("GET", "therapists", params={
            "name": f"eq.{name}",
            "select": "id",
        })
        if not existing:
            await sb_request("POST", "therapists", data={
                "name": name,
                "first_name": first,
                "last_name": last,
            })

    # Update upload metadata
    await sb_request("POST", "upload_metadata", data={
        "filename": file.filename,
        "uploaded_by": admin["id"],
        "transactions_count": len(transactions),
        "therapist_count": len(therapist_names),
        "date_range_start": date_min.strftime("%Y-%m-%d") if date_min else None,
        "date_range_end": date_max.strftime("%Y-%m-%d") if date_max else None,
    })

    date_range = ""
    if date_min and date_max:
        date_range = f"{date_min.strftime('%b %d, %Y')} — {date_max.strftime('%b %d, %Y')}"

    return {
        "status": "success",
        "transactions_count": len(transactions),
        "therapist_count": len(therapist_names),
        "date_range": date_range,
    }


@app.get("/api/settings/last-upload")
async def get_last_upload(user=Depends(verify_token)):
    """Get the most recent upload metadata."""
    results = await sb_request("GET", "upload_metadata", params={
        "select": "*",
        "order": "created_at.desc",
        "limit": "1",
    })
    if results:
        return {
            "uploaded_at": results[0].get("created_at"),
            "filename": results[0].get("filename"),
        }
    return None


# ── Analytics ────────────────────────────────────────────────────

@app.get("/api/analytics/summary")
async def analytics_summary(user=Depends(verify_token)):
    """
    Generate therapist analytics summary from transactions table.
    Preserves the existing LTV + engagement calculation logic.
    """
    # Fetch all transactions
    all_txns = []
    offset = 0
    page_size = 1000
    while True:
        page = await sb_request("GET", "transactions", params={
            "select": "*",
            "offset": str(offset),
            "limit": str(page_size),
        })
        if not page:
            break
        all_txns.extend(page)
        if len(page) < page_size:
            break
        offset += page_size

    if not all_txns:
        raise HTTPException(status_code=404, detail="No transaction data available")

    # ── Build analytics (mirrors existing Python logic) ──
    # Separate appointments and payments
    appointments = [t for t in all_txns if t.get("record_type", "").lower() in ("appointment", "charge")]
    payments = [t for t in all_txns if t.get("record_type", "").lower() == "payment"]

    # ── LTV Calculation ──
    # Get customer → provider mapping, expected insurance per customer
    customer_provider = {}  # customer → {provider: count}
    customer_expected_ins = {}  # customer → expected insurance amount
    customer_copay = {}  # customer → total copay payments

    for appt in appointments:
        name = appt.get("patient_name", "").strip()
        provider = appt.get("provider_name", "").strip()
        amount = abs(appt.get("amount", 0))

        if not name or not provider:
            continue

        if name not in customer_provider:
            customer_provider[name] = {}
        customer_provider[name][provider] = customer_provider[name].get(provider, 0) + 1

        # Expected insurance = billed amount (appointments typically represent billed amounts)
        customer_expected_ins[name] = customer_expected_ins.get(name, 0) + amount

    # Process payments
    attributed_insurance = {}  # customer → amount
    unattributed_insurance_total = 0.0

    for pmt in payments:
        name = pmt.get("patient_name", "").strip()
        amount = abs(pmt.get("amount", 0))
        payer = pmt.get("payer", "").strip().lower()

        is_patient_payment = "patient" in payer or "copay" in payer or "self" in payer or not payer
        is_insurance = not is_patient_payment

        if name and is_patient_payment:
            customer_copay[name] = customer_copay.get(name, 0) + amount
        elif name and is_insurance:
            attributed_insurance[name] = attributed_insurance.get(name, 0) + amount
        elif is_insurance:
            unattributed_insurance_total += amount

    # Proportional allocation of unattributed insurance
    total_expected = sum(customer_expected_ins.values()) or 1
    customer_allocated_ins = {}
    for name, expected in customer_expected_ins.items():
        share = expected / total_expected
        customer_allocated_ins[name] = share * unattributed_insurance_total

    # Calculate per-customer LTV
    all_customers = set(customer_copay.keys()) | set(attributed_insurance.keys()) | set(customer_expected_ins.keys())
    customer_ltv = {}
    for name in all_customers:
        copay = customer_copay.get(name, 0)
        ins_direct = attributed_insurance.get(name, 0)
        ins_allocated = customer_allocated_ins.get(name, 0)
        customer_ltv[name] = copay + ins_direct + ins_allocated

    # ── Per-therapist metrics ──
    # Determine primary therapist per customer
    customer_primary = {}
    for name, providers in customer_provider.items():
        customer_primary[name] = max(providers, key=providers.get)

    # APN handling: Tracey Nagle gets all clients, not just primary
    APN_NAMES = {"Tracey Nagle"}

    therapist_clients = {}  # therapist → set of clients
    therapist_appt_counts = {}  # therapist → {client: count}

    for name, providers in customer_provider.items():
        primary = customer_primary.get(name)
        if primary and primary not in APN_NAMES:
            if primary not in therapist_clients:
                therapist_clients[primary] = set()
                therapist_appt_counts[primary] = {}
            therapist_clients[primary].add(name)
            therapist_appt_counts[primary][name] = providers.get(primary, 0)

    # APN: all clients they've seen
    for name, providers in customer_provider.items():
        for apn in APN_NAMES:
            if apn in providers:
                if apn not in therapist_clients:
                    therapist_clients[apn] = set()
                    therapist_appt_counts[apn] = {}
                therapist_clients[apn].add(name)
                therapist_appt_counts[apn][name] = providers[apn]

    # Build therapist summary list
    therapist_list = []
    therapist_details = {}

    for therapist, clients in therapist_clients.items():
        client_count = len(clients)
        total_rev = sum(customer_ltv.get(c, 0) for c in clients)
        avg_ltv = total_rev / client_count if client_count else 0

        appts_per_client = [therapist_appt_counts[therapist].get(c, 0) for c in clients]
        avg_appts = sum(appts_per_client) / len(appts_per_client) if appts_per_client else 0
        total_appts = sum(appts_per_client)

        therapist_list.append({
            "name": therapist,
            "client_count": client_count,
            "ltv_contribution": round(total_rev, 2),
            "is_apn": therapist in APN_NAMES,
        })

        therapist_details[therapist] = {
            "client_count": client_count,
            "avg_ltv": round(avg_ltv, 2),
            "total_revenue": round(total_rev, 2),
            "avg_appointments": round(avg_appts, 1),
            "total_appointments": total_appts,
        }

    # Sort by LTV contribution descending
    therapist_list.sort(key=lambda x: x["ltv_contribution"], reverse=True)

    # Practice averages
    all_avg_ltv = list(therapist_details.values())
    practice_avg = {
        "avg_ltv": round(sum(d["avg_ltv"] for d in all_avg_ltv) / len(all_avg_ltv), 2) if all_avg_ltv else 0,
        "avg_appointments": round(sum(d["avg_appointments"] for d in all_avg_ltv) / len(all_avg_ltv), 1) if all_avg_ltv else 0,
    }

    return {
        "therapists": therapist_list,
        "therapist_details": therapist_details,
        "practice_avg": practice_avg,
        "total_customers": len(all_customers),
        "total_revenue": round(sum(customer_ltv.values()), 2),
    }


@app.get("/api/analytics/therapist/{user_id}")
async def therapist_analytics(user_id: str, user=Depends(verify_token)):
    """Get analytics for a specific therapist (by user_id)."""
    # Get the user's name to match against transactions
    users = await sb_request("GET", "users", params={
        "id": f"eq.{user_id}",
        "select": "first_name,last_name",
    })
    if not users:
        raise HTTPException(status_code=404, detail="User not found")

    therapist_name = f"{users[0]['first_name']} {users[0]['last_name']}"

    # Get full summary and extract this therapist's data
    try:
        # Use the summary endpoint logic
        summary_data = await analytics_summary.__wrapped__(user) if hasattr(analytics_summary, '__wrapped__') else None
    except Exception:
        summary_data = None

    if not summary_data:
        # Fallback: compute directly
        try:
            from starlette.requests import Request as _R
            # Re-use the summary function with a mock
            all_txns = []
            offset = 0
            while True:
                page = await sb_request("GET", "transactions", params={
                    "select": "*", "offset": str(offset), "limit": "1000",
                })
                if not page:
                    break
                all_txns.extend(page)
                if len(page) < 1000:
                    break
                offset += 1000

            if not all_txns:
                raise HTTPException(status_code=404, detail="No data")

            # Quick per-therapist calc
            my_appts = [t for t in all_txns if t.get("provider_name", "").strip() == therapist_name and t.get("record_type", "").lower() in ("appointment", "charge")]
            clients = set(t.get("patient_name", "").strip() for t in my_appts if t.get("patient_name"))
            client_count = len(clients)

            appts_by_client = {}
            for t in my_appts:
                c = t.get("patient_name", "").strip()
                if c:
                    appts_by_client[c] = appts_by_client.get(c, 0) + 1

            avg_appts = sum(appts_by_client.values()) / len(appts_by_client) if appts_by_client else 0

            return {
                "client_count": client_count,
                "avg_ltv": 0,
                "total_revenue": 0,
                "avg_appointments": round(avg_appts, 1),
                "practice_avg_ltv": 0,
                "practice_avg_appointments": 0,
            }

        except Exception as e:
            raise HTTPException(status_code=404, detail="No analytics data available")

    details = summary_data.get("therapist_details", {}).get(therapist_name, {})
    practice_avg = summary_data.get("practice_avg", {})

    return {
        **details,
        "practice_avg_ltv": practice_avg.get("avg_ltv", 0),
        "practice_avg_appointments": practice_avg.get("avg_appointments", 0),
    }


# ── Task Generation ──────────────────────────────────────────────

def compute_due_dates(template: dict, start: date, end: date) -> List[date]:
    """Return all due dates for a template between start and end (inclusive)."""
    schedule_type = template.get("schedule_type", "weekly")
    try:
        rule = json.loads(template.get("schedule_rule") or "{}")
    except Exception:
        rule = {}

    offset = int(template.get("default_due_offset_days") or 0)
    due_dates = []
    current = start

    while current <= end:
        include = False

        if schedule_type == "daily":
            every_n = int(rule.get("every_n_days", 1))
            # Include every N days from start
            delta = (current - start).days
            include = (delta % every_n == 0)

        elif schedule_type == "weekly":
            weekdays = rule.get("weekdays", [0, 1, 2, 3, 4])  # Mon–Fri default
            # Python weekday(): Mon=0, Sun=6
            include = current.weekday() in weekdays

        elif schedule_type == "monthly":
            day_of_month = int(rule.get("day_of_month", 1))
            include = current.day == day_of_month

        if include:
            due = current + timedelta(days=offset)
            due_dates.append(due)

        current += timedelta(days=1)

    return due_dates


@app.post("/api/tasks/generate")
async def generate_task_instances(days: int = 30, admin=Depends(require_admin)):
    """
    Generate task_instances for all active templates over the next `days` days.
    Safe to call repeatedly — unique constraint prevents duplicates.
    Admin-only.
    """
    today = date.today()
    window_end = today + timedelta(days=days)

    # Fetch all active templates
    templates = await sb_request("GET", "task_templates", params={
        "active": "eq.true",
        "select": "*",
    })

    if not templates:
        return {"status": "ok", "generated": 0, "skipped": 0, "message": "No active templates found"}

    generated = 0
    skipped = 0

    for tmpl in templates:
        due_dates = compute_due_dates(tmpl, today, window_end)

        for due in due_dates:
            instance = {
                "template_id": tmpl["id"],
                "title": tmpl["title"],
                "description": tmpl.get("description"),
                "tags": tmpl.get("tags", []),
                "priority": tmpl.get("priority", "medium"),
                "assigned_to_user_id": tmpl.get("assigned_to_user_id"),
                "assigned_to_role": tmpl.get("assigned_to_role"),
                "due_date": due.isoformat(),
                "status": "backlog",
            }

            try:
                await sb_request("POST", "task_instances", data=instance)
                generated += 1
            except HTTPException as e:
                # 409 Conflict = duplicate (unique constraint) — expected, skip silently
                if e.status_code in (409, 422, 400):
                    skipped += 1
                else:
                    logger.warning(f"Task instance insert failed: {e.detail}")
                    skipped += 1

    return {
        "status": "ok",
        "generated": generated,
        "skipped": skipped,
        "window_days": days,
        "templates_processed": len(templates),
    }


@app.get("/api/tasks/instances")
async def get_task_instances(user=Depends(verify_token)):
    """
    Get task instances for the current user.
    Admins see all; others see only their own (by user_id or role).
    """
    params = {
        "select": "*, task_templates(title, schedule_type)",
        "order": "due_date.asc",
    }

    if user.get("role") != "admin":
        # Filter to user's own instances or role-based ones
        params["or"] = f"(assigned_to_user_id.eq.{user['id']},assigned_to_role.eq.{user['role']})"

    instances = await sb_request("GET", "task_instances", params=params)
    return instances or []


@app.patch("/api/tasks/instances/{instance_id}")
async def update_task_instance(instance_id: str, request: Request, user=Depends(verify_token)):
    """Update a task instance status."""
    body = await request.json()
    allowed_fields = {"status", "completed_at"}
    update_data = {k: v for k, v in body.items() if k in allowed_fields}

    if "status" in update_data and update_data["status"] == "done":
        update_data["completed_at"] = datetime.utcnow().isoformat()
    elif "status" in update_data and update_data["status"] != "done":
        update_data["completed_at"] = None

    update_data["updated_at"] = datetime.utcnow().isoformat()

    result = await sb_request("PATCH", f"task_instances?id=eq.{instance_id}", data=update_data)
    return result


@app.get("/api/tasks/templates")
async def get_task_templates(admin=Depends(require_admin)):
    """Admin: list all task templates."""
    templates = await sb_request("GET", "task_templates", params={
        "select": "*",
        "order": "created_at.desc",
    })
    return templates or []


class TaskTemplateRequest(BaseModel):
    title: str
    description: Optional[str] = None
    tags: Optional[List[str]] = []
    priority: str = "medium"
    assigned_to_role: Optional[str] = None
    assigned_to_user_id: Optional[str] = None
    schedule_type: str = "weekly"
    schedule_rule: Optional[str] = "{}"
    timezone: str = "America/New_York"
    default_due_offset_days: int = 0
    active: bool = True


@app.post("/api/tasks/templates")
async def create_task_template(req: TaskTemplateRequest, admin=Depends(require_admin)):
    """Admin: create a new task template."""
    data = req.dict()
    data["created_by_user_id"] = admin["id"]
    result = await sb_request("POST", "task_templates", data=data)
    return result


@app.patch("/api/tasks/templates/{template_id}")
async def update_task_template(template_id: str, req: TaskTemplateRequest, admin=Depends(require_admin)):
    """Admin: update an existing task template."""
    data = req.dict()
    data["updated_at"] = datetime.utcnow().isoformat()
    result = await sb_request("PATCH", f"task_templates?id=eq.{template_id}", data=data)
    return result


@app.delete("/api/tasks/templates/{template_id}")
async def delete_task_template(template_id: str, admin=Depends(require_admin)):
    """Admin: soft-delete (deactivate) a task template."""
    result = await sb_request("PATCH", f"task_templates?id=eq.{template_id}", data={"active": False})
    return {"status": "deactivated", "id": template_id}


# ── Meeting Generation ────────────────────────────────────────────

def _nth_weekday_in_month(year: int, month: int, day_of_week: int, nth: int) -> Optional[date]:
    """
    Find the nth occurrence of a weekday in a given month.
    day_of_week: 0=Mon..6=Sun
    nth: 1-based (1=first, 2=second, ...) or -1 for last.
    Returns a date or None if invalid.
    """
    # calendar.monthcalendar returns weeks starting Monday by default
    cal = calendar.Calendar(firstweekday=0)  # Monday = 0
    month_days = cal.monthdayscalendar(year, month)

    if nth == -1:
        # Last occurrence: iterate weeks in reverse
        for week in reversed(month_days):
            d = week[day_of_week]
            if d != 0:
                return date(year, month, d)
        return None

    # Positive nth: count occurrences
    count = 0
    for week in month_days:
        d = week[day_of_week]
        if d != 0:
            count += 1
            if count == nth:
                return date(year, month, d)
    return None


def _is_last_weekday_of_month(d: date) -> bool:
    """Check if this date is the last occurrence of its weekday in the month."""
    # If adding 7 days pushes us to next month, it's the last one
    next_week = d + timedelta(days=7)
    return next_week.month != d.month


def compute_meeting_dates(template: dict, start: date, end: date) -> list:
    """Return all meeting dates for a template between start and end (inclusive)."""
    cadence = template.get("cadence", "weekly")
    rule = template.get("schedule_rule") or {}
    if isinstance(rule, str):
        try:
            rule = json.loads(rule)
        except Exception:
            rule = {}

    dates = []

    if cadence == "weekly":
        dow = int(rule.get("day_of_week", 0))
        skip_last = bool(rule.get("skip_last", False))
        current = start
        while current <= end:
            if current.weekday() == dow:
                if skip_last and _is_last_weekday_of_month(current):
                    pass  # Skip last occurrence of this weekday
                else:
                    dates.append(current)
            current += timedelta(days=1)

    elif cadence == "monthly":
        nth = int(rule.get("nth", 1))
        dow = int(rule.get("day_of_week", 0))
        # Iterate each month in range
        y, m = start.year, start.month
        while date(y, m, 1) <= end:
            d = _nth_weekday_in_month(y, m, dow, nth)
            if d and start <= d <= end:
                dates.append(d)
            # Next month
            m += 1
            if m > 12:
                m = 1
                y += 1

    elif cadence == "monthly_interval":
        every_n = int(rule.get("every_n_months", 1))
        anchor_str = rule.get("anchor", start.isoformat())
        weekday_rule = rule.get("weekday_rule")  # e.g. "first_monday"
        dom = int(rule.get("day_of_month", 1))
        try:
            anchor = date.fromisoformat(anchor_str)
        except Exception:
            anchor = start

        # Start from anchor, step by every_n_months
        y, m = anchor.year, anchor.month
        while True:
            d = None
            if weekday_rule == "first_monday":
                d = _nth_weekday_in_month(y, m, 0, 1)  # 1st Monday
            else:
                try:
                    d = date(y, m, min(dom, calendar.monthrange(y, m)[1]))
                except ValueError:
                    pass
            if d and d > end:
                break
            if d and start <= d <= end:
                dates.append(d)
            # Step forward
            m += every_n
            while m > 12:
                m -= 12
                y += 1

    elif cadence == "quarterly":
        # Two modes:
        # 1) "months": [4, 7, 10] — specific months with day 15 default
        # 2) "month_of_quarter" + "nth" + "day_of_week" — nth weekday in nth month of quarter
        specific_months = rule.get("months")
        if specific_months:
            day_of = int(rule.get("day", 15))
            for year in range(start.year, end.year + 1):
                for mo in specific_months:
                    try:
                        d = date(year, int(mo), min(day_of, calendar.monthrange(year, int(mo))[1]))
                        if start <= d <= end:
                            dates.append(d)
                    except ValueError:
                        pass
        else:
            moq = int(rule.get("month_of_quarter", 1))  # 1, 2, or 3
            nth = int(rule.get("nth", 1))
            dow = int(rule.get("day_of_week", 0))
            # Quarters: Q1=Jan-Mar, Q2=Apr-Jun, Q3=Jul-Sep, Q4=Oct-Dec
            quarter_starts = [1, 4, 7, 10]
            for qs in quarter_starts:
                target_month = qs + moq - 1
                for year in range(start.year, end.year + 1):
                    if target_month > 12:
                        continue
                    d = _nth_weekday_in_month(year, target_month, dow, nth)
                    if d and start <= d <= end:
                        dates.append(d)

    elif cadence == "yearly":
        mo = int(rule.get("month", 1))
        dy = int(rule.get("day", 1))
        for year in range(start.year, end.year + 1):
            try:
                d = date(year, mo, dy)
                if start <= d <= end:
                    dates.append(d)
            except ValueError:
                pass  # Invalid date (e.g., Feb 30)

    dates.sort()
    return dates


@app.post("/api/meetings/generate")
async def generate_meeting_instances(days: int = 120, admin=Depends(require_admin)):
    """
    Generate meeting_instances for all active templates over the next `days` days.
    Safe to call repeatedly — unique constraint prevents duplicates.
    Admin-only.
    """
    today = date.today()
    window_end = today + timedelta(days=days)

    templates = await sb_request("GET", "meeting_templates", params={
        "active": "eq.true",
        "select": "*",
    })

    if not templates:
        return {"status": "ok", "generated": 0, "skipped": 0, "message": "No active meeting templates found"}

    generated = 0
    skipped = 0

    for tmpl in templates:
        meeting_dates = compute_meeting_dates(tmpl, today, window_end)

        for md in meeting_dates:
            instance = {
                "template_id": tmpl["id"],
                "title": tmpl["title"],
                "meeting_date": md.isoformat(),
            }

            try:
                await sb_request("POST", "meeting_instances", data=instance)
                generated += 1
            except HTTPException as e:
                if e.status_code in (409, 422, 400):
                    skipped += 1
                else:
                    logger.warning(f"Meeting instance insert failed: {e.detail}")
                    skipped += 1

    return {
        "status": "ok",
        "generated": generated,
        "skipped": skipped,
        "window_days": days,
        "templates_processed": len(templates),
    }


@app.get("/api/meetings/instances")
async def get_meeting_instances(user=Depends(verify_token)):
    """
    Get upcoming meeting instances for the current user.
    Filters by audience_roles on the template (empty = all roles).
    Returns the next 20 meetings from today onward.
    """
    today_str = date.today().isoformat()
    user_role = user.get("role", "therapist")

    # Fetch upcoming instances with template info (service key to bypass RLS for join)
    instances = await sb_request("GET", "meeting_instances", params={
        "select": "*, meeting_templates(audience_roles, meeting_time)",
        "meeting_date": f"gte.{today_str}",
        "order": "meeting_date.asc",
        "limit": "50",
    })

    if not instances:
        return []

    # Filter by audience role client-side (since Supabase REST join filtering is limited)
    filtered = []
    for inst in instances:
        tmpl = inst.get("meeting_templates") or {}
        audience = tmpl.get("audience_roles") or []
        meeting_time = tmpl.get("meeting_time")
        # Empty audience = visible to all; admin sees everything
        if not audience or user_role == "admin" or user_role in audience:
            # Flatten meeting_time into the instance and remove nested template data
            inst.pop("meeting_templates", None)
            if meeting_time:
                inst["meeting_time"] = meeting_time
            filtered.append(inst)
        if len(filtered) >= 20:
            break

    return filtered


# ── Meeting Template CRUD ─────────────────────────────────────────

class MeetingTemplateRequest(BaseModel):
    title: str
    cadence: str = "weekly"
    schedule_rule: Optional[dict] = {}
    audience_roles: Optional[List[str]] = []
    meeting_time: Optional[str] = None
    active: bool = True


@app.get("/api/meetings/templates")
async def get_meeting_templates(admin=Depends(require_admin)):
    """Admin: list all meeting templates."""
    templates = await sb_request("GET", "meeting_templates", params={
        "select": "*",
        "order": "title.asc",
    })
    return templates or []


@app.post("/api/meetings/templates")
async def create_meeting_template(req: MeetingTemplateRequest, admin=Depends(require_admin)):
    """Admin: create a new meeting template."""
    data = req.dict()
    data["created_by_user_id"] = admin["id"]
    result = await sb_request("POST", "meeting_templates", data=data)
    return result


@app.patch("/api/meetings/templates/{template_id}")
async def update_meeting_template(template_id: str, req: MeetingTemplateRequest, admin=Depends(require_admin)):
    """Admin: update a meeting template."""
    data = req.dict()
    data["updated_at"] = datetime.utcnow().isoformat()
    result = await sb_request("PATCH", f"meeting_templates?id=eq.{template_id}", data=data)
    return result


@app.delete("/api/meetings/templates/{template_id}")
async def delete_meeting_template(template_id: str, admin=Depends(require_admin)):
    """Admin: deactivate a meeting template."""
    result = await sb_request("PATCH", f"meeting_templates?id=eq.{template_id}", data={"active": False})
    return {"status": "deactivated", "id": template_id}


# ── Meeting Instance CRUD ─────────────────────────────────────────

class MeetingInstanceRequest(BaseModel):
    title: str
    meeting_date: str
    template_id: Optional[str] = None


@app.delete("/api/meetings/instances/{instance_id}")
async def delete_meeting_instance(instance_id: str, admin=Depends(require_admin)):
    """Admin: delete a meeting instance."""
    await sb_request("DELETE", f"meeting_instances?id=eq.{instance_id}")
    return {"status": "deleted", "id": instance_id}


# ── Announcements CRUD ────────────────────────────────────────────

class AnnouncementRequest(BaseModel):
    title: str
    body: Optional[str] = None
    category: str = "general"
    audience_roles: Optional[List[str]] = []
    effective_date: str
    expiration_date: Optional[str] = None


@app.get("/api/announcements")
async def get_announcements(admin=Depends(require_admin)):
    """Admin: list all announcements (including expired)."""
    results = await sb_request("GET", "announcements", params={
        "select": "*",
        "order": "effective_date.desc",
    })
    return results or []


@app.post("/api/announcements")
async def create_announcement(req: AnnouncementRequest, admin=Depends(require_admin)):
    """Admin: create an announcement."""
    data = req.dict()
    data["created_by_user_id"] = admin["id"]
    result = await sb_request("POST", "announcements", data=data)
    return result


@app.patch("/api/announcements/{announcement_id}")
async def update_announcement(announcement_id: str, req: AnnouncementRequest, admin=Depends(require_admin)):
    """Admin: update an announcement."""
    data = req.dict()
    result = await sb_request("PATCH", f"announcements?id=eq.{announcement_id}", data=data)
    return result


@app.delete("/api/announcements/{announcement_id}")
async def delete_announcement(announcement_id: str, admin=Depends(require_admin)):
    """Admin: delete an announcement."""
    await sb_request("DELETE", f"announcements?id=eq.{announcement_id}")
    return {"status": "deleted", "id": announcement_id}


# ── Admin: List Users ──────────────────────────────────────────────

@app.get("/api/admin/users")
async def list_users(admin=Depends(require_admin)):
    """Admin: list all users."""
    users = await sb_request("GET", "users", params={
        "select": "*",
        "order": "last_name.asc",
    })
    return users or []


# ═══════════════════════════════════════════════════════════════════
# PAYROLL SYSTEM
# ═══════════════════════════════════════════════════════════════════

# ── Twilio Config ──────────────────────────────────────────────────
TWILIO_ACCOUNT_SID = os.environ.get("TWILIO_ACCOUNT_SID", "")
TWILIO_AUTH_TOKEN = os.environ.get("TWILIO_AUTH_TOKEN", "")
TWILIO_MESSAGING_SERVICE_SID = os.environ.get("TWILIO_MESSAGING_SERVICE_SID", "")

twilio_client = None

@app.on_event("startup")
async def init_twilio():
    global twilio_client
    if TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN:
        try:
            from twilio.rest import Client
            twilio_client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
            logger.info("Twilio client initialized")
        except ImportError:
            logger.warning("twilio package not installed — SMS disabled")
        except Exception as e:
            logger.warning(f"Twilio init failed: {e}")
    else:
        logger.info("Twilio credentials not configured — SMS disabled")


def send_sms(to_number: str, body: str):
    """Send SMS via Twilio Messaging Service."""
    if not twilio_client or not TWILIO_MESSAGING_SERVICE_SID:
        logger.info(f"SMS skipped (no Twilio): {to_number}")
        return None
    try:
        msg = twilio_client.messages.create(
            messaging_service_sid=TWILIO_MESSAGING_SERVICE_SID,
            to=to_number,
            body=body,
        )
        logger.info(f"SMS sent to {to_number}: {msg.sid}")
        return msg.sid
    except Exception as e:
        logger.error(f"SMS send failed: {e}")
        return None


# ── Rate Catalog ───────────────────────────────────────────────────

@app.get("/api/payroll/rate-catalog")
async def get_rate_catalog(user=Depends(verify_token)):
    """Get all rate types and bill rate defaults."""
    # Deactivate legacy duplicate ADOS types (parentheses versions replaced by dash versions)
    legacy_ados = ["ADOS Assessment (In Home)", "ADOS Assessment (In Office)", "ADOS In Home", "ADOS At Office"]
    for name in legacy_ados:
        await sb_request("PATCH", f"rate_types?name=eq.{name}&is_active=eq.true", data={"is_active": False})

    rate_types = await sb_request("GET", "rate_types", params={
        "select": "*",
        "is_active": "eq.true",
        "order": "sort_order.asc",
    })
    bill_defaults = await sb_request("GET", "bill_rate_defaults", params={
        "select": "*",
    })
    return {
        "rate_types": rate_types or [],
        "bill_rate_defaults": bill_defaults or [],
    }


class RateTypeRequest(BaseModel):
    name: str
    unit: str = "hourly"
    default_duration_minutes: Optional[int] = None
    default_bill_rate: Optional[float] = None


@app.post("/api/payroll/rate-types")
async def create_rate_type(req: RateTypeRequest, admin=Depends(require_admin)):
    """Admin: create a new rate type."""
    result = await sb_request("POST", "rate_types", data={
        "name": req.name,
        "unit": req.unit,
        "default_duration_minutes": req.default_duration_minutes,
    })
    if req.default_bill_rate and result:
        rt_id = result[0]["id"] if isinstance(result, list) else result["id"]
        await sb_request("POST", "bill_rate_defaults", data={
            "rate_type_id": rt_id,
            "default_bill_rate": req.default_bill_rate,
        })
    return result


@app.patch("/api/payroll/rate-types/{rate_type_id}")
async def update_rate_type(rate_type_id: str, req: RateTypeRequest, admin=Depends(require_admin)):
    """Admin: update a rate type."""
    await sb_request("PATCH", f"rate_types?id=eq.{rate_type_id}", data={
        "name": req.name,
        "unit": req.unit,
        "default_duration_minutes": req.default_duration_minutes,
        "updated_at": datetime.utcnow().isoformat(),
    })
    if req.default_bill_rate is not None:
        existing = await sb_request("GET", "bill_rate_defaults", params={
            "rate_type_id": f"eq.{rate_type_id}",
        })
        if existing:
            await sb_request("PATCH", f"bill_rate_defaults?rate_type_id=eq.{rate_type_id}", data={
                "default_bill_rate": req.default_bill_rate,
            })
        else:
            await sb_request("POST", "bill_rate_defaults", data={
                "rate_type_id": rate_type_id,
                "default_bill_rate": req.default_bill_rate,
            })
    return {"status": "updated"}


# ── User Pay Rates ─────────────────────────────────────────────────

@app.get("/api/payroll/user-pay-rates")
async def get_all_user_pay_rates(admin=Depends(require_admin)):
    """Admin: get all user pay rates."""
    return await sb_request("GET", "user_pay_rates", params={"select": "*"}) or []


class UserPayRatesRequest(BaseModel):
    rates: List[dict]


@app.post("/api/payroll/user-pay-rates/{user_id}")
async def set_user_pay_rates(user_id: str, req: UserPayRatesRequest, admin=Depends(require_admin)):
    """Admin: set pay rates for a user (upsert)."""
    saved = 0
    for rate in req.rates:
        rt_id = rate["rate_type_id"]
        pay_rate = rate["pay_rate"]
        # Check if existing rate exists for this user + rate_type (any date)
        existing = await sb_request("GET", "user_pay_rates", params={
            "user_id": f"eq.{user_id}",
            "rate_type_id": f"eq.{rt_id}",
            "select": "id",
            "limit": "1",
        })
        if existing:
            await sb_request("PATCH", f"user_pay_rates?user_id=eq.{user_id}&rate_type_id=eq.{rt_id}", data={
                "pay_rate": pay_rate,
                "updated_at": datetime.utcnow().isoformat(),
            })
        else:
            await sb_request("POST", "user_pay_rates", data={
                "user_id": user_id,
                "rate_type_id": rt_id,
                "pay_rate": pay_rate,
            })
        saved += 1

    return {"status": "saved", "count": saved}


# ── Pay Periods ────────────────────────────────────────────────────

@app.get("/api/payroll/pay-periods")
async def get_pay_periods(admin=Depends(require_admin)):
    """Admin: list all pay periods with recipient counts."""
    periods = await sb_request("GET", "pay_periods", params={
        "select": "*",
        "order": "start_date.desc",
    })
    if not periods:
        return []

    # Enrich with recipient counts
    for p in periods:
        recipients = await sb_request("GET", "pay_period_recipients", params={
            "pay_period_id": f"eq.{p['id']}",
            "select": "id,status",
        })
        p["recipient_count"] = len(recipients) if recipients else 0
        p["received_count"] = len([r for r in (recipients or []) if r["status"] in ("received", "approved", "exported")])

    return periods


class PayPeriodCreateRequest(BaseModel):
    period_type: str  # 'first_half' or 'second_half'


@app.post("/api/payroll/pay-periods")
async def create_pay_period(req: PayPeriodCreateRequest, admin=Depends(require_admin)):
    """Admin: create a new pay period."""
    today = date.today()

    if req.period_type == "first_half":
        start = date(today.year, today.month, 1)
        end = date(today.year, today.month, 15)
        due = end
    else:
        start = date(today.year, today.month, 16)
        _, last_day = calendar.monthrange(today.year, today.month)
        end = date(today.year, today.month, last_day)
        due = end

    label = f"{start.strftime('%b %d')} – {end.strftime('%b %d, %Y')}"

    result = await sb_request("POST", "pay_periods", data={
        "period_type": req.period_type,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "due_date": due.isoformat(),
        "status": "draft",
        "label": label,
        "created_by": admin["id"],
    })

    return result


@app.post("/api/payroll/pay-periods/{period_id}/open")
async def open_pay_period(period_id: str, admin=Depends(require_admin)):
    """
    Admin: open a pay period.
    - Auto-generates recipient list from active users
    - Sends initial email + SMS notifications
    """
    # Get the period
    periods = await sb_request("GET", "pay_periods", params={
        "id": f"eq.{period_id}",
        "select": "*",
    })
    if not periods:
        raise HTTPException(status_code=404, detail="Pay period not found")
    period = periods[0]

    if period["status"] != "draft":
        raise HTTPException(status_code=400, detail="Can only open draft periods")

    # Get active users (therapists, clinical_leaders, apn — not admin, front_desk)
    users = await sb_request("GET", "users", params={
        "is_active": "eq.true",
        "select": "id,first_name,last_name,email,phone_number,sms_enabled,role",
    })

    payroll_roles = {"therapist", "clinical_leader", "apn"}
    eligible = [u for u in (users or []) if u.get("role") in payroll_roles]
    logger.info(f"Opening pay period {period_id}: {len(users or [])} total users, {len(eligible)} eligible")

    # Create recipients
    created = 0
    sms_sent = 0
    sms_skipped = 0
    for user in eligible:
        try:
            recipient = await sb_request("POST", "pay_period_recipients", data={
                "pay_period_id": period_id,
                "user_id": user["id"],
                "status": "sent",
            })
            created += 1
            logger.info(f"Created recipient for {user.get('first_name')} {user.get('last_name')} ({user['id']})")

            # Send notification
            name = user.get("first_name", "")
            msg = f"Hi {name}! Your BestLife invoice for {period['label']} is now open. Please submit by {period['due_date']}."

            # SMS
            phone = user.get("phone_number")
            sms_on = user.get("sms_enabled")
            logger.info(f"  SMS check: phone={phone}, sms_enabled={sms_on}, twilio_client={'yes' if twilio_client else 'no'}")
            if phone and sms_on:
                sid = send_sms(phone, msg)
                if sid:
                    sms_sent += 1
                    if recipient:
                        rid = recipient[0]["id"] if isinstance(recipient, list) else recipient["id"]
                        await sb_request("POST", "reminder_log", data={
                            "recipient_id": rid,
                            "channel": "sms",
                            "status": "sent",
                        })
                else:
                    sms_skipped += 1
                    logger.warning(f"  SMS send returned None for {phone}")
            else:
                sms_skipped += 1

        except Exception as e:
            logger.warning(f"Failed to create recipient for {user['id']}: {e}")

    logger.info(f"Pay period opened: {created} recipients, {sms_sent} SMS sent, {sms_skipped} SMS skipped")

    # Update period status
    await sb_request("PATCH", f"pay_periods?id=eq.{period_id}", data={
        "status": "open",
        "opened_at": datetime.utcnow().isoformat(),
    })

    # Audit log
    await sb_request("POST", "audit_log", data={
        "action": "pay_period_opened",
        "entity_type": "pay_period",
        "entity_id": period_id,
        "user_id": admin["id"],
        "details": {"recipients_created": created},
    })

    return {"status": "opened", "recipients_created": created}


@app.post("/api/payroll/pay-periods/{period_id}/close")
async def close_pay_period(period_id: str, admin=Depends(require_admin)):
    """Admin: close a pay period."""
    await sb_request("PATCH", f"pay_periods?id=eq.{period_id}", data={
        "status": "closed",
        "closed_at": datetime.utcnow().isoformat(),
    })
    return {"status": "closed"}


@app.post("/api/payroll/pay-periods/{period_id}/reopen")
async def reopen_pay_period(period_id: str, admin=Depends(require_admin)):
    """Admin: reopen a closed pay period (undo close)."""
    periods = await sb_request("GET", "pay_periods", params={
        "id": f"eq.{period_id}", "select": "id,status",
    })
    if not periods:
        raise HTTPException(status_code=404, detail="Pay period not found")
    if periods[0]["status"] != "closed":
        raise HTTPException(status_code=400, detail="Only closed periods can be reopened")
    await sb_request("PATCH", f"pay_periods?id=eq.{period_id}", data={
        "status": "open",
        "closed_at": None,
    })
    return {"status": "reopened"}


@app.delete("/api/payroll/pay-periods/{period_id}")
async def delete_pay_period(period_id: str, admin=Depends(require_admin)):
    """Admin: permanently delete a pay period and all associated data."""
    periods = await sb_request("GET", "pay_periods", params={
        "id": f"eq.{period_id}", "select": "id,status",
    })
    if not periods:
        raise HTTPException(status_code=404, detail="Pay period not found")
    # Delete time_entries first (FK to pay_period_recipients)
    await sb_request("DELETE", f"time_entries?pay_period_id=eq.{period_id}")
    # Delete rollup data
    await sb_request("DELETE", f"rollup_pay_period?pay_period_id=eq.{period_id}")
    # Delete recipients (FK to pay_periods)
    await sb_request("DELETE", f"pay_period_recipients?pay_period_id=eq.{period_id}")
    # Delete the period itself
    await sb_request("DELETE", f"pay_periods?id=eq.{period_id}")

    # Audit log
    await sb_request("POST", "audit_log", data={
        "action": "pay_period_deleted",
        "entity_type": "pay_period",
        "entity_id": period_id,
        "user_id": admin["id"],
        "details": {"status": periods[0].get("status")},
    })

    return {"status": "deleted"}


@app.get("/api/payroll/pay-periods/{period_id}/recipients")
async def get_period_recipients(period_id: str, admin=Depends(require_admin)):
    """Admin: list all recipients for a pay period with their draft tokens."""
    # Note: pay_period_recipients has TWO FKs to users (user_id, approved_by).
    # We must disambiguate by specifying the FK column in the join hint:
    #   users!user_id(...)  tells PostgREST which FK to use.
    recipients = await sb_request("GET", "pay_period_recipients", params={
        "pay_period_id": f"eq.{period_id}",
        "select": "id,user_id,status,draft_token,submit_token,submitted_at,users!user_id(first_name,last_name,email)",
        "order": "created_at.asc",
    })

    logger.info(f"Recipients raw for period {period_id}: {recipients}")

    result = []
    for r in (recipients or []):
        user = r.pop("users", {}) or {}
        result.append({
            **r,
            "user_name": f"{user.get('first_name', '')} {user.get('last_name', '')}".strip(),
            "user_email": user.get("email", ""),
        })

    return result


# ── Approval Queue ─────────────────────────────────────────────────

@app.get("/api/payroll/approval-queue")
async def get_approval_queue(status: str = "received", admin=Depends(require_admin)):
    """Admin: get recipients filtered by status."""
    params = {
        "select": "*, users!user_id(first_name, last_name, role), pay_periods(start_date, end_date, label)",
        "order": "updated_at.desc",
    }
    if status != "all":
        params["status"] = f"eq.{status}"

    recipients = await sb_request("GET", "pay_period_recipients", params=params)

    result = []
    for r in (recipients or []):
        user = r.pop("users", {}) or {}
        period = r.pop("pay_periods", {}) or {}
        r["first_name"] = user.get("first_name", "")
        r["last_name"] = user.get("last_name", "")
        r["user_name"] = f"{r['first_name']} {r['last_name']}".strip()
        r["period_start"] = period.get("start_date")
        r["period_end"] = period.get("end_date")
        r["period_label"] = period.get("label")
        result.append(r)

    return result


class ApproveRequest(BaseModel):
    overrides: Optional[dict] = None

class RejectRequest(BaseModel):
    reason: str

class ZeroHoursRequest(BaseModel):
    reason: str


@app.post("/api/payroll/recipients/{recipient_id}/approve")
async def approve_recipient(recipient_id: str, req: ApproveRequest, admin=Depends(require_admin)):
    """
    Admin: approve a submission.
    - Writes immutable time_entries from invoice_data
    - Calculates est_bill and est_pay using rate tables
    - Updates rollups
    """
    # Get recipient
    recipients = await sb_request("GET", "pay_period_recipients", params={
        "id": f"eq.{recipient_id}",
        "select": "*",
    })
    if not recipients:
        raise HTTPException(status_code=404, detail="Recipient not found")
    recipient = recipients[0]

    if recipient["status"] not in ("received",):
        raise HTTPException(status_code=400, detail="Can only approve received submissions")

    invoice_data = recipient.get("invoice_data") or {}
    user_id = recipient["user_id"]
    period_id = recipient["pay_period_id"]

    # Get user's pay rates
    pay_rates = await sb_request("GET", "user_pay_rates", params={
        "user_id": f"eq.{user_id}",
        "select": "*, rate_types(name, unit)",
    })
    pay_rate_map = {}
    for pr in (pay_rates or []):
        rt = pr.get("rate_types") or {}
        pay_rate_map[rt.get("name", "")] = {
            "rate_type_id": pr["rate_type_id"],
            "pay_rate": float(pr["pay_rate"]),
            "unit": rt.get("unit", "hourly"),
        }

    # Get bill rate defaults
    bill_defaults = await sb_request("GET", "bill_rate_defaults", params={
        "select": "*, rate_types(name)",
    })
    bill_rate_map = {}
    for bd in (bill_defaults or []):
        rt = bd.get("rate_types") or {}
        bill_rate_map[rt.get("name", "")] = float(bd["default_bill_rate"])

    # ── Parse structured invoice_data ──
    # Format: { iic: { CODE: [{cyber_initials, date, hours},...] }, op: { sessions: [{client_initials, date, cancel_fee?},...] },
    #   sbys: [{date, hours},...], ados: [{client_initials, location, id_number, date},...],
    #   admin: [{date, hours},...], supervision: { individual: [...], group: [...] },
    #   sick_leave: { date, hours, reason }, pto: { hours }, notes: "..." }
    total_bill = 0.0
    total_pay = 0.0
    total_hours = 0.0
    total_sessions = 0

    # Helper: write a time entry and accumulate totals
    async def write_entry(rate_name, qty, initials=None, duration=None, notes=None):
        nonlocal total_bill, total_pay, total_hours, total_sessions
        if qty == 0:
            return
        rate_info = pay_rate_map.get(rate_name, {})
        rate_type_id = rate_info.get("rate_type_id")
        pay_rate = rate_info.get("pay_rate", 0)
        bill_rate = bill_rate_map.get(rate_name, 0)
        unit = rate_info.get("unit", "hourly")
        est_bill = bill_rate * qty
        est_pay = pay_rate * qty
        if rate_type_id:
            await sb_request("POST", "time_entries", data={
                "recipient_id": recipient_id, "user_id": user_id, "pay_period_id": period_id,
                "rate_type_id": rate_type_id, "quantity": qty,
                "duration_minutes": duration, "client_initials": initials,
                "est_bill_amount": round(est_bill, 2), "est_pay_amount": round(est_pay, 2),
                "notes": notes, "locked": True,
            })
        total_bill += est_bill
        total_pay += est_pay
        if unit == "hourly":
            total_hours += qty
        else:
            total_sessions += int(qty)

    # IIC sessions
    iic = invoice_data.get("iic") or {}
    for code, entries in iic.items():
        if not isinstance(entries, list):
            continue
        # Map IIC code to rate type name
        iic_rate_names = {
            "IICLC-H0036TJU1": "IIC LPC/LCSW",
            "IICMA-H0036TJU2": "IIC LAC/LSW",
            "BA-H2014TJ": "Behavioral Assistant",
        }
        rate_name = iic_rate_names.get(code, code)
        for entry in entries:
            hrs = float(entry.get("hours") or 0)
            await write_entry(rate_name, hrs, initials=entry.get("cyber_initials"))

    # OP sessions (each session = 1 hour, cancellations also = 1 hour)
    op = invoice_data.get("op") or {}
    op_sessions = op.get("sessions") or []
    for entry in op_sessions:
        is_cancel = entry.get("cancel_fee")
        rate_name = "OP Cancellation" if is_cancel else "OP Session"
        await write_entry(rate_name, 1.0, initials=entry.get("client_initials"))

    # SBYS
    sbys = invoice_data.get("sbys") or []
    for entry in sbys:
        hrs = float(entry.get("hours") or 0)
        await write_entry("SBYS", hrs)

    # ADOS (each assessment = 3 hours toward time worked)
    ados = invoice_data.get("ados") or []
    for entry in ados:
        await write_entry("ADOS", 3.0, initials=entry.get("client_initials"),
                          notes=f"{entry.get('location','')} ID:{entry.get('id_number','')}")

    # Admin
    admin_entries = invoice_data.get("admin") or []
    for entry in admin_entries:
        hrs = float(entry.get("hours") or 0)
        await write_entry("Administration", hrs)

    # Supervision (individual + group, each session = 1 hour)
    sup = invoice_data.get("supervision") or {}
    for s in (sup.get("individual") or []):
        await write_entry("Individual Supervision", 1.0, notes=f"Supervisor: {s.get('supervisor_name','')}")
    for s in (sup.get("group") or []):
        await write_entry("Group Supervision", 1.0, notes=f"Supervisees: {','.join(s.get('supervisee_names',[]))}")

    # Sick Leave (only if admin approved)
    sick = invoice_data.get("sick_leave") or {}
    sick_hrs = float(sick.get("hours") or 0)
    if sick_hrs > 0:
        await write_entry("Sick Leave", sick_hrs, notes=sick.get("reason"))

    # PTO
    pto = invoice_data.get("pto") or {}
    pto_hrs = float(pto.get("hours") or 0)
    if pto_hrs > 0:
        await write_entry("PTO", pto_hrs)

    # Update recipient status
    await sb_request("PATCH", f"pay_period_recipients?id=eq.{recipient_id}", data={
        "status": "approved",
        "approved_at": datetime.utcnow().isoformat(),
        "approved_by": admin["id"],
        "admin_override_data": req.overrides,
        "updated_at": datetime.utcnow().isoformat(),
    })

    # Update/Insert rollup_pay_period
    existing_rollup = await sb_request("GET", "rollup_pay_period", params={
        "pay_period_id": f"eq.{period_id}",
        "user_id": f"eq.{user_id}",
    })
    rollup_data = {
        "pay_period_id": period_id,
        "user_id": user_id,
        "total_hours": round(total_hours, 2),
        "total_sessions": total_sessions,
        "est_bill_total": round(total_bill, 2),
        "est_pay_total": round(total_pay, 2),
        "margin": round(total_bill - total_pay, 2),
        "updated_at": datetime.utcnow().isoformat(),
    }
    if existing_rollup:
        await sb_request("PATCH", f"rollup_pay_period?pay_period_id=eq.{period_id}&user_id=eq.{user_id}", data=rollup_data)
    else:
        await sb_request("POST", "rollup_pay_period", data=rollup_data)

    # Update rollup_monthly based on service_date month
    # For simplicity, use the pay period's start_date month
    period_data = await sb_request("GET", "pay_periods", params={"id": f"eq.{period_id}", "select": "start_date"})
    if period_data:
        month_year = period_data[0]["start_date"][:7]  # 'YYYY-MM'
        existing_monthly = await sb_request("GET", "rollup_monthly", params={
            "user_id": f"eq.{user_id}",
            "month_year": f"eq.{month_year}",
        })
        monthly_data = {
            "user_id": user_id,
            "month_year": month_year,
            "total_hours": round(total_hours, 2),
            "total_sessions": total_sessions,
            "est_bill_total": round(total_bill, 2),
            "est_pay_total": round(total_pay, 2),
            "margin": round(total_bill - total_pay, 2),
            "updated_at": datetime.utcnow().isoformat(),
        }
        if existing_monthly:
            # Accumulate
            old = existing_monthly[0]
            monthly_data["total_hours"] = round(float(old.get("total_hours", 0)) + total_hours, 2)
            monthly_data["total_sessions"] = int(old.get("total_sessions", 0)) + total_sessions
            monthly_data["est_bill_total"] = round(float(old.get("est_bill_total", 0)) + total_bill, 2)
            monthly_data["est_pay_total"] = round(float(old.get("est_pay_total", 0)) + total_pay, 2)
            monthly_data["margin"] = round(monthly_data["est_bill_total"] - monthly_data["est_pay_total"], 2)
            await sb_request("PATCH", f"rollup_monthly?user_id=eq.{user_id}&month_year=eq.{month_year}", data=monthly_data)
        else:
            await sb_request("POST", "rollup_monthly", data=monthly_data)

    # Audit log
    await sb_request("POST", "audit_log", data={
        "action": "recipient_approved",
        "entity_type": "pay_period_recipient",
        "entity_id": recipient_id,
        "user_id": admin["id"],
        "details": {"total_bill": round(total_bill, 2), "total_pay": round(total_pay, 2)},
    })

    return {"status": "approved", "total_bill": round(total_bill, 2), "total_pay": round(total_pay, 2)}


@app.post("/api/payroll/recipients/{recipient_id}/reject")
async def reject_recipient(recipient_id: str, req: RejectRequest, admin=Depends(require_admin)):
    """Admin: reject a submission."""
    await sb_request("PATCH", f"pay_period_recipients?id=eq.{recipient_id}", data={
        "status": "rejected",
        "rejection_reason": req.reason,
        "approved_by": admin["id"],
        "updated_at": datetime.utcnow().isoformat(),
    })
    return {"status": "rejected"}


class AdminNoteRequest(BaseModel):
    note: str


@app.post("/api/payroll/recipients/{recipient_id}/admin-note")
async def add_admin_note(recipient_id: str, req: AdminNoteRequest, admin=Depends(require_admin)):
    """Admin: add a note/question to a recipient's submission."""
    recipients = await sb_request("GET", "pay_period_recipients", params={
        "id": f"eq.{recipient_id}", "select": "id,admin_notes",
    })
    if not recipients:
        raise HTTPException(status_code=404, detail="Recipient not found")

    existing_notes = recipients[0].get("admin_notes") or []
    new_note = {
        "text": req.note,
        "by": f"{admin['first_name']} {admin['last_name']}",
        "by_id": admin["id"],
        "at": datetime.utcnow().isoformat(),
    }
    existing_notes.append(new_note)

    await sb_request("PATCH", f"pay_period_recipients?id=eq.{recipient_id}", data={
        "admin_notes": existing_notes,
        "updated_at": datetime.utcnow().isoformat(),
    })
    return {"status": "note_added", "notes": existing_notes}


class UpdateLineItemsRequest(BaseModel):
    invoice_data: dict


@app.patch("/api/payroll/recipients/{recipient_id}/invoice-data")
async def update_invoice_data(recipient_id: str, req: UpdateLineItemsRequest, admin=Depends(require_admin)):
    """Admin: edit line items on a submitted invoice before approval."""
    recipients = await sb_request("GET", "pay_period_recipients", params={
        "id": f"eq.{recipient_id}", "select": "id,status",
    })
    if not recipients:
        raise HTTPException(status_code=404, detail="Recipient not found")
    if recipients[0]["status"] not in ("received",):
        raise HTTPException(status_code=400, detail="Can only edit received (pending) submissions")

    await sb_request("PATCH", f"pay_period_recipients?id=eq.{recipient_id}", data={
        "invoice_data": req.invoice_data,
        "updated_at": datetime.utcnow().isoformat(),
    })

    await sb_request("POST", "audit_log", data={
        "action": "invoice_data_edited",
        "entity_type": "pay_period_recipient",
        "entity_id": recipient_id,
        "user_id": admin["id"],
    })

    return {"status": "updated"}


@app.post("/api/payroll/recipients/{recipient_id}/zero-hours")
async def zero_hours_recipient(recipient_id: str, req: ZeroHoursRequest, admin=Depends(require_admin)):
    """Admin: mark as zero hours with reason."""
    await sb_request("PATCH", f"pay_period_recipients?id=eq.{recipient_id}", data={
        "status": "zero_hours",
        "zero_hours_reason": req.reason,
        "approved_by": admin["id"],
        "updated_at": datetime.utcnow().isoformat(),
    })
    return {"status": "zero_hours"}


# ── Export Batches ─────────────────────────────────────────────────

@app.get("/api/payroll/export-batches")
async def get_export_batches(admin=Depends(require_admin)):
    """Admin: list all export batches + count of exportable."""
    batches = await sb_request("GET", "export_batches", params={
        "select": "*",
        "order": "created_at.desc",
    })
    # Count approved (not yet exported)
    approved = await sb_request("GET", "pay_period_recipients", params={
        "status": "eq.approved",
        "select": "id",
    })
    return {
        "batches": batches or [],
        "exportable_count": len(approved) if approved else 0,
    }


@app.post("/api/payroll/export-batches/generate")
async def generate_export_batch(admin=Depends(require_admin)):
    """
    Admin: generate an export batch from all approved (not yet exported) recipients.
    Prevents double export.
    """
    # Get all approved recipients
    recipients = await sb_request("GET", "pay_period_recipients", params={
        "status": "eq.approved",
        "select": "*, users(first_name, last_name, email), pay_periods(start_date, end_date, label)",
    })

    if not recipients:
        raise HTTPException(status_code=400, detail="No approved recipients to export")

    # Build CSV
    csv_lines = ["Name,Email,Pay Period,Total Hours,Est Bill,Est Pay,Margin"]
    total_pay = 0.0
    recipient_ids = []

    for r in recipients:
        user = r.get("users") or {}
        period = r.get("pay_periods") or {}
        name = f"{user.get('first_name', '')} {user.get('last_name', '')}".strip()
        email = user.get("email", "")
        period_label = period.get("label", f"{period.get('start_date', '')} - {period.get('end_date', '')}")

        # Get rollup for this recipient
        rollup = await sb_request("GET", "rollup_pay_period", params={
            "pay_period_id": f"eq.{r['pay_period_id']}",
            "user_id": f"eq.{r['user_id']}",
            "select": "*",
        })
        rp = rollup[0] if rollup else {}

        hours = rp.get("total_hours", 0)
        bill = rp.get("est_bill_total", 0)
        pay = rp.get("est_pay_total", 0)
        margin = rp.get("margin", 0)

        csv_lines.append(f"{name},{email},{period_label},{hours},{bill},{pay},{margin}")
        total_pay += float(pay)
        recipient_ids.append(r["id"])

    csv_text = "\n".join(csv_lines)
    filename = f"payroll-export-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.csv"

    # Create batch record
    batch = await sb_request("POST", "export_batches", data={
        "record_count": len(recipient_ids),
        "total_pay": round(total_pay, 2),
        "csv_text": csv_text,
        "exported_by": admin["id"],
        "label": filename,
    })

    # Mark recipients as exported
    for rid in recipient_ids:
        await sb_request("PATCH", f"pay_period_recipients?id=eq.{rid}", data={
            "status": "exported",
            "updated_at": datetime.utcnow().isoformat(),
        })

    # Audit log
    await sb_request("POST", "audit_log", data={
        "action": "export_batch_created",
        "entity_type": "export_batch",
        "entity_id": batch[0]["id"] if isinstance(batch, list) else batch.get("id"),
        "user_id": admin["id"],
        "details": {"record_count": len(recipient_ids), "total_pay": round(total_pay, 2)},
    })

    return {"status": "exported", "csv_text": csv_text, "filename": filename, "record_count": len(recipient_ids)}


@app.get("/api/payroll/export-batches/{batch_id}/download")
async def download_export_batch(batch_id: str, admin=Depends(require_admin)):
    """Admin: download a previously generated batch."""
    batches = await sb_request("GET", "export_batches", params={
        "id": f"eq.{batch_id}",
        "select": "*",
    })
    if not batches:
        raise HTTPException(status_code=404, detail="Batch not found")
    b = batches[0]
    return {"csv_text": b.get("csv_text", ""), "filename": b.get("label", "export.csv")}


# ── Public Invoice Flow (no auth required) ─────────────────────────

@app.get("/api/public/invoice/{draft_token}")
async def get_public_invoice(draft_token: str):
    """
    Public: get invoice form for a recipient via draft token.
    Returns rate types, user role, supervisees (for clinical leaders), and draft data.
    """
    recipients = await sb_request("GET", "pay_period_recipients", params={
        "draft_token": f"eq.{draft_token}",
        "select": "*, users!user_id(id, first_name, last_name, role), pay_periods(start_date, end_date, label, due_date, status)",
    })
    if not recipients:
        raise HTTPException(status_code=404, detail="Invoice link not found or expired")

    r = recipients[0]
    period = r.get("pay_periods") or {}

    if period.get("status") != "open":
        raise HTTPException(status_code=400, detail="This pay period is no longer accepting submissions")

    if r["status"] in ("received", "approved", "exported"):
        return {"already_submitted": True, "submitted_at": r.get("submitted_at")}

    # Get rate types
    rate_types = await sb_request("GET", "rate_types", params={
        "is_active": "eq.true",
        "select": "*",
        "order": "sort_order.asc",
    })

    user = r.get("users") or {}
    user_role = user.get("role", "therapist")

    # If clinical leader, fetch their supervisees for the supervision section
    supervisees = []
    if user_role == "clinical_leader":
        team = await sb_request("GET", "users", params={
            "clinical_supervisor_id": f"eq.{user.get('id', r['user_id'])}",
            "is_active": "eq.true",
            "select": "id,first_name,last_name",
            "order": "first_name.asc",
        })
        supervisees = [
            {"id": s["id"], "name": f"{s.get('first_name', '')} {s.get('last_name', '')}".strip()}
            for s in (team or [])
        ]

    return {
        "recipient_id": r["id"],
        "user_name": f"{user.get('first_name', '')} {user.get('last_name', '')}".strip(),
        "user_role": user_role,
        "period_label": period.get("label", ""),
        "due_date": period.get("due_date"),
        "draft_data": r.get("invoice_data"),
        "submit_token": r.get("submit_token"),
        "rate_types": rate_types or [],
        "supervisees": supervisees,
    }


class DraftSaveRequest(BaseModel):
    invoice_data: dict


@app.post("/api/public/invoice/{draft_token}/save-draft")
async def save_draft(draft_token: str, req: DraftSaveRequest):
    """Public: save draft invoice data."""
    recipients = await sb_request("GET", "pay_period_recipients", params={
        "draft_token": f"eq.{draft_token}",
        "select": "id,status",
    })
    if not recipients:
        raise HTTPException(status_code=404, detail="Not found")

    r = recipients[0]
    if r["status"] in ("received", "approved", "exported"):
        raise HTTPException(status_code=400, detail="Already submitted")

    await sb_request("PATCH", f"pay_period_recipients?id=eq.{r['id']}", data={
        "invoice_data": req.invoice_data,
        "updated_at": datetime.utcnow().isoformat(),
    })

    return {"status": "draft_saved"}


class SubmitRequest(BaseModel):
    submit_token: str
    invoice_data: dict


@app.post("/api/public/invoice/{draft_token}/submit")
async def submit_invoice(draft_token: str, req: SubmitRequest):
    """
    Public: submit the invoice (single submit only).
    Validates submit_token, saves final invoice_data, marks as received.
    """
    recipients = await sb_request("GET", "pay_period_recipients", params={
        "draft_token": f"eq.{draft_token}",
        "select": "id,status,submit_token",
    })
    if not recipients:
        raise HTTPException(status_code=404, detail="Not found")

    r = recipients[0]
    if r["status"] in ("received", "approved", "exported"):
        raise HTTPException(status_code=400, detail="Already submitted — single submit only")

    if str(r.get("submit_token")) != req.submit_token:
        raise HTTPException(status_code=403, detail="Invalid submit token")

    await sb_request("PATCH", f"pay_period_recipients?id=eq.{r['id']}", data={
        "invoice_data": req.invoice_data,
        "status": "received",
        "submitted_at": datetime.utcnow().isoformat(),
        "updated_at": datetime.utcnow().isoformat(),
    })

    return {"status": "submitted"}


# ── Reminders ──────────────────────────────────────────────────────

@app.post("/api/payroll/send-reminders")
async def send_reminders(admin=Depends(require_admin)):
    """
    Admin: send reminders for open pay periods.
    Cadence: on open, day 3, day before due, morning of due.
    Only sends if status is 'sent' (not yet received).
    """
    # Get open periods
    periods = await sb_request("GET", "pay_periods", params={
        "status": "eq.open",
        "select": "*",
    })

    if not periods:
        return {"status": "no_open_periods", "sent": 0}

    sent_count = 0
    today = date.today()

    for period in periods:
        due = date.fromisoformat(period["due_date"])
        opened = period.get("opened_at")
        if opened:
            opened_date = datetime.fromisoformat(opened.replace("Z", "+00:00")).date()
        else:
            opened_date = today

        days_since_open = (today - opened_date).days
        days_until_due = (due - today).days

        # Check if today is a reminder day
        should_send = (
            days_since_open == 3 or
            days_until_due == 1 or
            days_until_due == 0
        )

        if not should_send:
            continue

        # Get recipients who haven't submitted
        recipients = await sb_request("GET", "pay_period_recipients", params={
            "pay_period_id": f"eq.{period['id']}",
            "status": "eq.sent",
            "select": "*, users(first_name, phone_number, sms_enabled)",
        })

        for r in (recipients or []):
            user = r.get("users") or {}
            name = user.get("first_name", "")
            phone = user.get("phone_number")
            sms_ok = user.get("sms_enabled", False)

            if phone and sms_ok:
                if days_until_due == 0:
                    msg = f"Hi {name} — your BestLife invoice for {period['label']} is due TODAY! Please submit ASAP."
                elif days_until_due == 1:
                    msg = f"Hi {name} — reminder: your BestLife invoice for {period['label']} is due tomorrow."
                else:
                    msg = f"Hi {name} — friendly reminder to submit your BestLife invoice for {period['label']} (due {period['due_date']})."

                sid = send_sms(phone, msg)
                if sid:
                    await sb_request("POST", "reminder_log", data={
                        "recipient_id": r["id"],
                        "channel": "sms",
                        "status": "sent",
                    })
                    await sb_request("PATCH", f"pay_period_recipients?id=eq.{r['id']}", data={
                        "reminder_count": (r.get("reminder_count") or 0) + 1,
                        "last_reminder_at": datetime.utcnow().isoformat(),
                    })
                    sent_count += 1

    return {"status": "ok", "sent": sent_count}


# ── Analytics (Rollup-Based) ───────────────────────────────────────

@app.get("/api/analytics/hours-margin")
async def analytics_hours_margin(view: str = "pay_period", user=Depends(verify_token)):
    """Analytics: Hours & Margin from rollups."""
    if view == "monthly":
        rollups = await sb_request("GET", "rollup_monthly", params={
            "select": "*, users(first_name, last_name)",
            "order": "month_year.desc",
        })
    else:
        rollups = await sb_request("GET", "rollup_pay_period", params={
            "select": "*, users(first_name, last_name), pay_periods(label)",
            "order": "updated_at.desc",
        })

    rows = []
    for r in (rollups or []):
        u = r.get("users") or {}
        rows.append({
            "user_name": f"{u.get('first_name', '')} {u.get('last_name', '')}".strip(),
            "total_hours": float(r.get("total_hours", 0)),
            "est_bill": float(r.get("est_bill_total", 0)),
            "est_pay": float(r.get("est_pay_total", 0)),
            "margin": float(r.get("margin", 0)),
            "period": r.get("pay_periods", {}).get("label") if view != "monthly" else r.get("month_year"),
        })

    return {"rows": rows}


# ── Billing Summary (new) ──────────────────────────────────────────

# Revenue-generating service types (split by license/location)
REVENUE_TYPES = ["IIC-LC", "IIC-MA", "IIC-BA", "OP", "SBYS", "ADOS In Home", "ADOS At Office", "APN 30 Min", "APN Intake"]
# Non-revenue (still tracked for hours)
NON_REVENUE_TYPES = ["PTO", "Sick Leave"]
SERVICE_TYPES = REVENUE_TYPES + NON_REVENUE_TYPES

# Map IIC billing codes to split categories
IIC_CODE_MAP = {
    "IICLC-H0036TJU1": "IIC-LC",
    "IICMA-H0036TJU2": "IIC-MA",
    "BA-H2014TJ": "IIC-BA",
}

# Map service type keys → rate_types.name in the DB (for bill rate lookup/save)
SERVICE_TO_RATE_NAME = {
    "IIC-LC": "IIC-LC",
    "IIC-MA": "IIC-MA",
    "IIC-BA": "IIC-BA",
    "OP": "OP-LC Session",
    "SBYS": "SBYS",
    "ADOS In Home": "ADOS Assessment - In Home",
    "ADOS At Office": "ADOS Assessment - At Office",
    "APN 30 Min": "APN Session (30)",
    "APN Intake": "APN Intake (60)",
}
# Reverse map
RATE_NAME_TO_SERVICE = {v: k for k, v in SERVICE_TO_RATE_NAME.items()}

def _extract_service_hours(invoice_data: dict) -> dict:
    """Extract hours per service type from structured invoice_data."""
    hours = {st: 0.0 for st in SERVICE_TYPES}
    # Also track ADOS count (assessments, not hours)
    counts = {"ADOS In Home": 0, "ADOS At Office": 0}
    if not invoice_data:
        return hours, counts

    # IIC — split by license code
    iic = invoice_data.get("iic") or {}
    for code, entries in iic.items():
        if isinstance(entries, list):
            stype = IIC_CODE_MAP.get(code, "IIC-LC")  # default to LC
            for e in entries:
                hours[stype] += float(e.get("hours") or 0)

    # OP (each session = 1 hour, cancellations also = 1 hour)
    op = invoice_data.get("op") or {}
    for s in (op.get("sessions") or []):
        hours["OP"] += 1.0

    # SBYS
    for e in (invoice_data.get("sbys") or []):
        hours["SBYS"] += float(e.get("hours") or 0)

    # ADOS — split by location (each assessment = 3 hours toward time worked)
    for e in (invoice_data.get("ados") or []):
        loc = (e.get("location") or "").lower()
        if "office" in loc:
            hours["ADOS At Office"] += 3.0
            counts["ADOS At Office"] += 1
        else:
            hours["ADOS In Home"] += 3.0
            counts["ADOS In Home"] += 1

    # APN — split by duration
    for e in (invoice_data.get("apn") or []):
        mins = float(e.get("duration_minutes") or e.get("minutes") or 30)
        hrs_val = float(e.get("hours") or (mins / 60))
        if mins >= 50:  # intake (60 min)
            hours["APN Intake"] += hrs_val
        else:
            hours["APN 30 Min"] += hrs_val

    # PTO
    pto = invoice_data.get("pto") or {}
    hours["PTO"] += float(pto.get("hours") or 0)

    # Sick Leave
    sick = invoice_data.get("sick_leave") or {}
    hours["Sick Leave"] += float(sick.get("hours") or 0)

    return hours, counts


@app.get("/api/analytics/billing-summary")
async def billing_summary(admin=Depends(require_admin)):
    """
    Billing Summary: returns all pay periods (open and closed) with per-service-type breakdowns.
    Closed periods aggregate approved recipients only; open periods include all submitted invoices.
    """
    # Get all pay periods (open and closed), most recent first
    periods = await sb_request("GET", "pay_periods", params={
        "select": "id, label, start_date, end_date, status",
        "order": "start_date.desc",
    })
    if not periods:
        periods = []

    # Get bill rate config for projected revenue
    bill_rates = await sb_request("GET", "bill_rate_defaults", params={
        "select": "*, rate_types(name)",
    })
    bill_rate_map = {}
    for br in (bill_rates or []):
        rt = br.get("rate_types") or {}
        name = rt.get("name", "")
        bill_rate_map[name] = float(br.get("default_bill_rate", 0))

    # Build service→bill_rate map using explicit name mapping
    service_bill_rates = {}
    for st in REVENUE_TYPES:
        rate_name = SERVICE_TO_RATE_NAME.get(st, st)
        rate = bill_rate_map.get(rate_name, 0)
        if not rate:
            # Fallback: try exact match on service key itself
            rate = bill_rate_map.get(st, 0)
        service_bill_rates[st] = rate

    # Get pay rates for all users (needed for per-service pay calculation)
    all_pay_rates_list = await sb_request("GET", "user_pay_rates", params={
        "select": "user_id, pay_rate, rate_types(name)",
    })
    user_pay_map_list = {}
    for pr in (all_pay_rates_list or []):
        uid = pr["user_id"]
        rt = pr.get("rate_types") or {}
        rname = rt.get("name", "")
        if uid not in user_pay_map_list:
            user_pay_map_list[uid] = {}
        user_pay_map_list[uid][rname] = float(pr["pay_rate"])

    period_summaries = []

    for period in periods:
        pid = period["id"]
        period_status = period.get("status", "open")
        # For closed periods: only approved recipients; for open periods: all submitted invoices
        recipient_params = {
            "pay_period_id": f"eq.{pid}",
            "select": "id, user_id, invoice_data, status, users!user_id(first_name, last_name)",
        }
        if period_status == "closed":
            recipient_params["status"] = "eq.approved"
        recipients = await sb_request("GET", "pay_period_recipients", params=recipient_params)

        # Aggregate by service type AND compute per-service pay
        service_totals = {st: 0.0 for st in SERVICE_TYPES}
        service_pay = {st: 0.0 for st in SERVICE_TYPES}
        ados_counts = {"ADOS In Home": 0, "ADOS At Office": 0}
        for r in (recipients or []):
            uid = r["user_id"]
            hours, counts = _extract_service_hours(r.get("invoice_data") or {})
            user_rates = user_pay_map_list.get(uid, {})
            for st, h in hours.items():
                service_totals[st] += h
                # Compute pay for this user's hours in this service type
                rate_name = SERVICE_TO_RATE_NAME.get(st, st)
                pay_rate = user_rates.get(rate_name, 0)
                if not pay_rate:
                    for rn, rv in user_rates.items():
                        if st.lower() in rn.lower() or rn.lower() in st.lower():
                            pay_rate = rv
                            break
                if st.startswith("ADOS"):
                    # ADOS pay = assessments × session rate
                    num_a = counts.get(st, 0)
                    service_pay[st] += num_a * pay_rate
                else:
                    service_pay[st] += h * pay_rate
            for k, v in counts.items():
                ados_counts[k] = ados_counts.get(k, 0) + v

        # Calculate revenue, pay, profit per service type
        service_breakdown = []
        grand_hours = 0.0
        grand_revenue = 0.0
        grand_pay = 0.0
        for st in SERVICE_TYPES:
            hrs = service_totals[st]
            if hrs == 0:
                continue
            bill_rate = service_bill_rates.get(st, 0)
            # ADOS revenue = bill_rate × number of assessments (not hours)
            if st.startswith("ADOS") and st in REVENUE_TYPES:
                num_assessments = ados_counts.get(st, 0)
                revenue = num_assessments * bill_rate
            elif st in REVENUE_TYPES:
                revenue = hrs * bill_rate
            else:
                revenue = 0
            pay = service_pay.get(st, 0)
            grand_hours += hrs
            grand_revenue += revenue
            grand_pay += pay
            entry = {
                "service": st,
                "hours": round(hrs, 2),
                "revenue": round(revenue, 2),
                "pay": round(pay, 2),
                "bill_rate": bill_rate,
            }
            # Add assessment count for ADOS
            if st.startswith("ADOS"):
                entry["assessments"] = ados_counts.get(st, 0)
            service_breakdown.append(entry)

        period_summaries.append({
            "id": pid,
            "label": period.get("label", ""),
            "start_date": period["start_date"],
            "end_date": period["end_date"],
            "services": service_breakdown,
            "total_hours": round(grand_hours, 2),
            "total_revenue": round(grand_revenue, 2),
            "total_pay": round(grand_pay, 2),
            "total_profit": round(grand_revenue - grand_pay, 2),
            "margin_pct": round((grand_revenue - grand_pay) / grand_revenue * 100, 1) if grand_revenue > 0 else 0,
        })

    # Build monthly summaries (group periods by month)
    monthly = {}
    for ps in period_summaries:
        month_key = ps["start_date"][:7]  # YYYY-MM
        if month_key not in monthly:
            monthly[month_key] = {
                "month": month_key,
                "services": {st: {"hours": 0.0, "revenue": 0.0, "pay": 0.0, "assessments": 0} for st in SERVICE_TYPES},
                "total_hours": 0, "total_revenue": 0, "total_pay": 0,
            }
        for svc in ps["services"]:
            monthly[month_key]["services"][svc["service"]]["hours"] += svc["hours"]
            monthly[month_key]["services"][svc["service"]]["revenue"] += svc.get("revenue", 0)
            monthly[month_key]["services"][svc["service"]]["pay"] += svc.get("pay", 0)
            monthly[month_key]["services"][svc["service"]]["assessments"] += svc.get("assessments", 0)
        monthly[month_key]["total_hours"] += ps["total_hours"]
        monthly[month_key]["total_revenue"] += ps["total_revenue"]
        monthly[month_key]["total_pay"] += ps["total_pay"]

    monthly_list = []
    for mk, mv in sorted(monthly.items(), reverse=True):
        svc_list = []
        for st in SERVICE_TYPES:
            sd = mv["services"][st]
            hrs = sd["hours"]
            if hrs == 0:
                continue
            entry = {
                "service": st,
                "hours": round(hrs, 2),
                "revenue": round(sd["revenue"], 2),
                "pay": round(sd["pay"], 2),
            }
            if st.startswith("ADOS"):
                entry["assessments"] = sd["assessments"]
            svc_list.append(entry)
        profit = mv["total_revenue"] - mv["total_pay"]
        monthly_list.append({
            "month": mk,
            "services": svc_list,
            "total_hours": round(mv["total_hours"], 2),
            "total_revenue": round(mv["total_revenue"], 2),
            "total_pay": round(mv["total_pay"], 2),
            "total_profit": round(profit, 2),
            "margin_pct": round(profit / mv["total_revenue"] * 100, 1) if mv["total_revenue"] > 0 else 0,
        })

    return {
        "periods": period_summaries,
        "monthly": monthly_list,
        "bill_rates": service_bill_rates,
    }


@app.get("/api/analytics/billing-summary/{period_id}")
async def billing_summary_detail(period_id: str, admin=Depends(require_admin)):
    """
    Detail view for a single pay period: per-therapist breakdown by service type,
    mirroring the billing master sheet layout.
    """
    # Get the period
    periods = await sb_request("GET", "pay_periods", params={
        "id": f"eq.{period_id}", "select": "id, label, start_date, end_date",
    })
    if not periods:
        raise HTTPException(status_code=404, detail="Pay period not found")
    period = periods[0]

    # Get approved recipients
    recipients = await sb_request("GET", "pay_period_recipients", params={
        "pay_period_id": f"eq.{period_id}",
        "status": "eq.approved",
        "select": "id, user_id, invoice_data, users!user_id(first_name, last_name)",
    })

    # Get bill rates
    bill_rates = await sb_request("GET", "bill_rate_defaults", params={
        "select": "*, rate_types(name)",
    })
    bill_rate_map = {}
    for br in (bill_rates or []):
        rt = br.get("rate_types") or {}
        bill_rate_map[rt.get("name", "")] = float(br.get("default_bill_rate", 0))

    service_bill_rates = {}
    for st in REVENUE_TYPES:
        rate_name = SERVICE_TO_RATE_NAME.get(st, st)
        rate = bill_rate_map.get(rate_name, 0)
        if not rate:
            rate = bill_rate_map.get(st, 0)
        service_bill_rates[st] = rate

    # Get pay rates for all users
    all_pay_rates = await sb_request("GET", "user_pay_rates", params={
        "select": "user_id, pay_rate, rate_types(name)",
    })
    # user_id → { rate_name: pay_rate }
    user_pay_map = {}
    for pr in (all_pay_rates or []):
        uid = pr["user_id"]
        rt = pr.get("rate_types") or {}
        rname = rt.get("name", "")
        if uid not in user_pay_map:
            user_pay_map[uid] = {}
        user_pay_map[uid][rname] = float(pr["pay_rate"])

    # Build per-service-type per-therapist detail
    service_detail = {st: [] for st in SERVICE_TYPES}
    service_totals_map = {st: {"hours": 0, "revenue": 0, "pay": 0} for st in SERVICE_TYPES}

    for r in (recipients or []):
        u = r.get("users") or {}
        name = f"{u.get('first_name', '')} {u.get('last_name', '')}".strip()
        uid = r["user_id"]
        inv = r.get("invoice_data") or {}
        hours_map, counts_map = _extract_service_hours(inv)
        user_rates = user_pay_map.get(uid, {})

        for st in SERVICE_TYPES:
            hrs = hours_map[st]
            if hrs == 0:
                continue
            bill_rate = service_bill_rates.get(st, 0)
            # ADOS revenue = bill_rate × assessments (not hours)
            if st.startswith("ADOS") and st in REVENUE_TYPES:
                num_a = counts_map.get(st, 0)
                revenue = num_a * bill_rate
            elif st in REVENUE_TYPES:
                revenue = hrs * bill_rate
            else:
                revenue = 0
            # Find the user's pay rate using rate name mapping
            rate_name = SERVICE_TO_RATE_NAME.get(st, st)
            pay_rate = user_rates.get(rate_name, 0)
            if not pay_rate:
                # Try partial match
                for rn, rv in user_rates.items():
                    if st.lower() in rn.lower() or rn.lower() in st.lower():
                        pay_rate = rv
                        break
            if not pay_rate:
                # Fallback: generic hourly
                pay_rate = user_rates.get("hourly", user_rates.get("Hourly", 0))
                if not pay_rate and user_rates:
                    pay_rate = list(user_rates.values())[0]
            # ADOS pay = assessments × session rate (not hours × rate)
            if st.startswith("ADOS"):
                num_a = counts_map.get(st, 0)
                pay = num_a * pay_rate
            else:
                pay = hrs * pay_rate
            profit = revenue - pay
            margin = (profit / revenue * 100) if revenue > 0 else 0

            service_detail[st].append({
                "name": name,
                "hours": round(hrs, 2),
                "revenue": round(revenue, 2),
                "pay": round(pay, 2),
                "profit": round(profit, 2),
                "margin": round(margin, 1),
            })
            service_totals_map[st]["hours"] += hrs
            service_totals_map[st]["revenue"] += revenue
            service_totals_map[st]["pay"] += pay

    # Build response
    sections = []
    for st in SERVICE_TYPES:
        if not service_detail[st]:
            continue
        t = service_totals_map[st]
        profit = t["revenue"] - t["pay"]
        sections.append({
            "service": st,
            "rows": service_detail[st],
            "total_hours": round(t["hours"], 2),
            "total_revenue": round(t["revenue"], 2),
            "total_pay": round(t["pay"], 2),
            "total_profit": round(profit, 2),
            "total_margin": round(profit / t["revenue"] * 100, 1) if t["revenue"] > 0 else 0,
        })

    # Grand totals
    gh = sum(s["total_hours"] for s in sections)
    gr = sum(s["total_revenue"] for s in sections)
    gp = sum(s["total_pay"] for s in sections)

    return {
        "period": period,
        "sections": sections,
        "grand_total": {
            "hours": round(gh, 2),
            "revenue": round(gr, 2),
            "pay": round(gp, 2),
            "profit": round(gr - gp, 2),
            "margin": round((gr - gp) / gr * 100, 1) if gr > 0 else 0,
        },
        "bill_rates": service_bill_rates,
    }


@app.patch("/api/analytics/billing-rates")
async def update_billing_rates(req: dict = Body(...), admin=Depends(require_admin)):
    """Admin: update projected revenue rates per service type."""
    # req format: { "IIC-LC": 123.45, "OP": 115, ... }
    rate_types_list = await sb_request("GET", "rate_types", params={"select": "id, name"})
    rt_map = {rt["name"]: rt["id"] for rt in (rate_types_list or [])}

    for service_key, rate_value in req.items():
        # Map service key to DB rate_type name
        db_name = SERVICE_TO_RATE_NAME.get(service_key, service_key)
        rt_id = rt_map.get(db_name)
        if not rt_id:
            # Also try the service key directly
            rt_id = rt_map.get(service_key)
        if not rt_id:
            # Create the rate type
            result = await sb_request("POST", "rate_types", data={
                "name": db_name, "unit": "session" if "ADOS" in service_key else "hourly",
            })
            if result:
                rt_id = result[0]["id"] if isinstance(result, list) else result["id"]
                rt_map[db_name] = rt_id

        if rt_id:
            existing = await sb_request("GET", "bill_rate_defaults", params={
                "rate_type_id": f"eq.{rt_id}",
            })
            if existing:
                await sb_request("PATCH", f"bill_rate_defaults?rate_type_id=eq.{rt_id}", data={
                    "default_bill_rate": float(rate_value),
                })
            else:
                await sb_request("POST", "bill_rate_defaults", data={
                    "rate_type_id": rt_id,
                    "default_bill_rate": float(rate_value),
                })

    return {"status": "updated"}


@app.post("/api/admin/migrate-rate-types")
async def migrate_rate_types(admin=Depends(require_admin)):
    """One-time migration: update rate_types for v5 schema changes.
    Deactivates old types, adds IIC-LC/MA/BA, splits OP, renames ADOS."""

    # Step 1: Deactivate types to remove
    deactivate_names = [
        "IIC", "APN 30 Min", "ADOS Assessment (In Home)", "ADOS Assessment (In Office)",
        "ADOS In Home", "ADOS At Office",
        "Other (Hourly)", "Other (Day)", "APN Other (Custom)",
    ]
    for name in deactivate_names:
        await sb_request("PATCH", f"rate_types?name=eq.{name}", data={"is_active": False})

    # Step 2: Add new IIC split types
    new_types = [
        {"name": "IIC-LC", "unit": "hourly", "sort_order": 1},
        {"name": "IIC-MA", "unit": "hourly", "sort_order": 2},
        {"name": "IIC-BA", "unit": "hourly", "sort_order": 3},
    ]
    for nt in new_types:
        existing = await sb_request("GET", "rate_types", params={"name": f"eq.{nt['name']}"})
        if existing:
            await sb_request("PATCH", f"rate_types?name=eq.{nt['name']}", data={"sort_order": nt["sort_order"], "is_active": True})
        else:
            await sb_request("POST", "rate_types", data=nt)

    # Step 3: Rename OP Session → OP-LC Session, add OP-MA Session
    existing_op = await sb_request("GET", "rate_types", params={"name": "eq.OP Session"})
    if existing_op:
        await sb_request("PATCH", f"rate_types?name=eq.OP Session", data={"name": "OP-LC Session", "sort_order": 4})

    existing_op_ma = await sb_request("GET", "rate_types", params={"name": "eq.OP-MA Session"})
    if not existing_op_ma:
        await sb_request("POST", "rate_types", data={
            "name": "OP-MA Session", "unit": "hourly", "default_duration_minutes": 60, "sort_order": 5,
        })
    else:
        await sb_request("PATCH", f"rate_types?name=eq.OP-MA Session", data={"sort_order": 5, "is_active": True})

    # Step 4: Add new ADOS types (session-based)
    ados_types = [
        {"name": "ADOS Assessment - In Home", "unit": "session", "sort_order": 8},
        {"name": "ADOS Assessment - At Office", "unit": "session", "sort_order": 9},
    ]
    for at in ados_types:
        existing = await sb_request("GET", "rate_types", params={"name": f"eq.{at['name']}"})
        if existing:
            await sb_request("PATCH", f"rate_types?name=eq.{at['name']}", data={"sort_order": at["sort_order"], "is_active": True})
        else:
            await sb_request("POST", "rate_types", data=at)

    # Step 5: Reorder remaining types
    reorder = {
        "SBYS": 6, "Administration": 7, "APN Session (30)": 10, "APN Intake (60)": 11,
        "PTO": 12, "Sick Leave": 13, "Community Event (Day)": 14, "OP Cancellation": 15,
    }
    for name, order in reorder.items():
        await sb_request("PATCH", f"rate_types?name=eq.{name}", data={"sort_order": order})

    return {"status": "migration_complete", "message": "Rate types updated successfully"}


# ── Performance Tracking Thresholds ──
PERF_THRESHOLDS = {
    "full_time": {"monthly": 80, "per_period": 40},
    "part_time": {"monthly": 40, "per_period": 20},
    "1099": {"monthly": 20, "per_period": 10},
}

# Rate type names → display column mapping for performance grid
# Must cover ALL possible names (pre-migration, post-migration, and invoice-submitted names)
RATE_TO_PERF_COL = {
    # IIC (all variants)
    "IIC": "iic", "IIC-LC": "iic", "IIC-MA": "iic", "IIC-BA": "iic",
    "IIC LPC/LCSW": "iic", "IIC LAC/LSW": "iic", "Behavioral Assistant": "iic",
    # OP (all variants)
    "OP Session": "op", "OP-LC Session": "op", "OP-MA Session": "op", "OP Cancellation": "op",
    # SBYS
    "SBYS": "sbys",
    # ADOS (all variants)
    "ADOS": "ados",
    "ADOS Assessment (In Home)": "ados", "ADOS Assessment (In Office)": "ados",
    "ADOS Assessment - In Home": "ados", "ADOS Assessment - At Office": "ados",
    "ADOS In Home": "ados", "ADOS At Office": "ados",
    # Time off
    "PTO": "pto", "Sick Leave": "sick",
    # Admin
    "Administration": "admin", "Community Event (Day)": "admin",
    # APN
    "APN Session (30)": "iic", "APN Intake (60)": "iic",
}


def _quarter_for_month(month_str: str):
    """Given 'YYYY-MM', return ('YYYY-Q1', ['YYYY-01','YYYY-02','YYYY-03'])."""
    y, m = int(month_str[:4]), int(month_str[5:7])
    q = (m - 1) // 3 + 1
    start_m = (q - 1) * 3 + 1
    months = [f"{y}-{start_m + i:02d}" for i in range(3)]
    return f"{y}-Q{q}", months


@app.get("/api/analytics/performance")
async def analytics_performance(
    timeframe: str = "monthly",
    period: str = "",
    user=Depends(verify_token),
):
    """Performance tracking: per-service hours, thresholds, status badges, grouped by leader."""
    profile = user
    caller_role = profile.get("role", "therapist")
    caller_id = profile.get("id")

    today = date.today()

    # ── 1. Determine date range + period info ──
    if timeframe == "pay_period" and period:
        # Specific pay period by ID
        pp = await sb_request("GET", "pay_periods", params={"id": f"eq.{period}", "select": "id,start_date,end_date,label"})
        if not pp:
            raise HTTPException(status_code=404, detail="Pay period not found")
        start_date = pp[0]["start_date"]
        end_date = pp[0]["end_date"]
        period_label = pp[0].get("label", f"{start_date} – {end_date}")
    elif timeframe == "quarterly":
        # period = "YYYY-QN"
        if period and "-Q" in period:
            y = int(period[:4])
            q = int(period[-1])
        else:
            y, q = today.year, (today.month - 1) // 3 + 1
        sm = (q - 1) * 3 + 1
        start_date = f"{y}-{sm:02d}-01"
        em = sm + 2
        last_day = calendar.monthrange(y, em)[1]
        end_date = f"{y}-{em:02d}-{last_day}"
        period_label = f"Q{q} {y}"
        period = f"{y}-Q{q}"
    elif timeframe == "yearly":
        y = int(period) if period and period.isdigit() else today.year
        start_date = f"{y}-01-01"
        end_date = f"{y}-12-31"
        period_label = str(y)
        period = str(y)
    else:
        # monthly (default)
        timeframe = "monthly"
        if period and len(period) == 7:
            y, m = int(period[:4]), int(period[5:7])
        else:
            y, m = today.year, today.month
            period = f"{y}-{m:02d}"
        start_date = f"{y}-{m:02d}-01"
        last_day = calendar.monthrange(y, m)[1]
        end_date = f"{y}-{m:02d}-{last_day}"
        months_names = ['January','February','March','April','May','June','July','August','September','October','November','December']
        period_label = f"{months_names[m-1]} {y}"

    # ── 2. Build available periods for dropdown ──
    all_periods = await sb_request("GET", "pay_periods", params={
        "select": "id,label,start_date,end_date,status",
        "status": "eq.closed",
        "order": "start_date.desc",
    }) or []

    available_periods = []
    if timeframe == "pay_period":
        for p in all_periods:
            available_periods.append({"value": p["id"], "label": p.get("label", p["start_date"])})
        if not period and available_periods:
            period = available_periods[0]["value"]
            pp = await sb_request("GET", "pay_periods", params={"id": f"eq.{period}", "select": "id,start_date,end_date,label"})
            if pp:
                start_date = pp[0]["start_date"]
                end_date = pp[0]["end_date"]
                period_label = pp[0].get("label", start_date)
    elif timeframe == "monthly":
        seen = set()
        for p in all_periods:
            mk = p["start_date"][:7]
            if mk not in seen:
                seen.add(mk)
                y2, m2 = int(mk[:4]), int(mk[5:7])
                months_names = ['January','February','March','April','May','June','July','August','September','October','November','December']
                available_periods.append({"value": mk, "label": f"{months_names[m2-1]} {y2}"})
    elif timeframe == "quarterly":
        seen = set()
        for p in all_periods:
            y2, m2 = int(p["start_date"][:4]), int(p["start_date"][5:7])
            qk = f"{y2}-Q{(m2-1)//3+1}"
            if qk not in seen:
                seen.add(qk)
                available_periods.append({"value": qk, "label": qk.replace("-", " ")})
    elif timeframe == "yearly":
        seen = set()
        for p in all_periods:
            yk = p["start_date"][:4]
            if yk not in seen:
                seen.add(yk)
                available_periods.append({"value": yk, "label": yk})

    # ── 3. Fetch users (trackable roles only) ──
    trackable_roles = ["therapist", "clinical_leader", "apn"]
    all_users = await sb_request("GET", "users", params={
        "is_active": "eq.true",
        "select": "id,first_name,last_name,role,employment_status,clinical_supervisor_id",
    }) or []
    # Filter to trackable roles
    all_users = [u for u in all_users if u.get("role") in trackable_roles]

    # Role-based visibility
    if caller_role == "admin":
        visible_users = all_users
    elif caller_role == "clinical_leader":
        visible_users = [u for u in all_users if u["id"] == caller_id or u.get("clinical_supervisor_id") == caller_id]
    else:
        visible_users = [u for u in all_users if u["id"] == caller_id]

    user_map = {u["id"]: u for u in visible_users}

    # ── 4. Find pay periods that overlap the date range ──
    range_periods = await sb_request("GET", "pay_periods", params={
        "select": "id,start_date,end_date",
        "start_date": f"lte.{end_date}",
        "end_date": f"gte.{start_date}",
        "status": "eq.closed",
    }) or []
    period_ids = [p["id"] for p in range_periods]
    num_periods = max(len(period_ids), 1)

    # ── 5. Fetch time_entries for those periods, joined to rate_types ──
    user_service_hours = {uid: {"iic": 0, "op": 0, "sbys": 0, "ados": 0, "sick": 0, "pto": 0, "admin": 0, "total": 0} for uid in user_map}

    if period_ids:
        for pid in period_ids:
            entries = await sb_request("GET", "time_entries", params={
                "pay_period_id": f"eq.{pid}",
                "select": "user_id,quantity,client_initials,rate_types(name)",
            }) or []
            for e in entries:
                uid = e["user_id"]
                if uid not in user_service_hours:
                    continue
                rt = e.get("rate_types") or {}
                rname = rt.get("name", "")
                qty = float(e.get("quantity", 0))
                col = RATE_TO_PERF_COL.get(rname, None)
                # time_entries already store qty in hours (ADOS entries stored as 3.0 per assessment)
                hrs = qty
                if col and col in user_service_hours[uid]:
                    user_service_hours[uid][col] += hrs
                user_service_hours[uid]["total"] += hrs

    # ── 6. Fetch therapist_capacity ──
    caps = await sb_request("GET", "therapist_capacity", params={"select": "user_id,iic_capacity,op_capacity"}) or []
    cap_map = {c["user_id"]: c for c in caps}

    # ── 7. Compute status badges (quarterly accountability) ──
    # For monthly view: determine which quarter this month belongs to
    # Then check each month in the quarter for compliance
    if timeframe == "monthly":
        current_month = period  # "YYYY-MM"
    elif timeframe == "pay_period" and period_ids:
        current_month = start_date[:7]
    else:
        current_month = f"{today.year}-{today.month:02d}"

    _, quarter_months = _quarter_for_month(current_month)

    # Get total hours per user per month in this quarter (from rollup_monthly)
    quarter_compliance = {uid: {} for uid in user_map}  # uid → {month: total_hours}
    for qm in quarter_months:
        rollups = await sb_request("GET", "rollup_monthly", params={
            "month_year": f"eq.{qm}",
            "select": "user_id,total_hours",
        }) or []
        for r in rollups:
            uid = r["user_id"]
            if uid in quarter_compliance:
                quarter_compliance[uid][qm] = float(r.get("total_hours", 0))

    def compute_status(uid, emp_status):
        threshold = PERF_THRESHOLDS.get(emp_status, PERF_THRESHOLDS["full_time"])["monthly"]
        trend = []
        months_below = 0
        for qm in quarter_months:
            hrs = quarter_compliance.get(uid, {}).get(qm, 0)
            met = hrs >= threshold
            trend.append(met)
            if not met and qm <= current_month:
                months_below += 1
        # Current month compliance
        current_hrs = quarter_compliance.get(uid, {}).get(current_month, 0)
        current_met = current_hrs >= threshold
        if current_met:
            status = "on_track"
        elif months_below >= 2:
            status = "action_required"
        else:
            status = "warning"
        return status, trend

    # ── 8. Group by clinical_supervisor_id ──
    # Clinical leaders are always their own group header (never unassigned)
    groups_map = {}  # leader_id → list of therapist dicts
    for uid, u in user_map.items():
        leader_id = u.get("clinical_supervisor_id")
        # Clinical leaders group under themselves (they ARE the leader)
        if u.get("role") == "clinical_leader":
            leader_id = uid
        emp = u.get("employment_status", "full_time")
        hrs = user_service_hours.get(uid, {})
        cap = cap_map.get(uid, {})
        status, trend = compute_status(uid, emp)

        iic_cap = int(cap.get("iic_capacity", 0) or 0)
        op_cap = int(cap.get("op_capacity", 0) or 0)
        total_cap = iic_cap + op_cap
        util_pct = None  # Placeholder — caseload not implemented yet

        therapist_row = {
            "user_id": uid,
            "name": f"{u.get('first_name', '')} {u.get('last_name', '')}".strip(),
            "role": u.get("role", ""),
            "is_leader": u.get("role") == "clinical_leader",
            "employment_status": emp,
            "status": status,
            "quarter_trend": trend,
            "iic": round(hrs.get("iic", 0), 1),
            "op": round(hrs.get("op", 0), 1),
            "sbys": round(hrs.get("sbys", 0), 1),
            "ados": round(hrs.get("ados", 0), 1),
            "sick": round(hrs.get("sick", 0), 1),
            "pto": round(hrs.get("pto", 0), 1),
            "admin_hours": round(hrs.get("admin", 0), 1),
            "total_hours": round(hrs.get("total", 0), 1),
            "avg_per_period": round(hrs.get("total", 0) / num_periods, 1),
            "caseload_iic": None,  # Placeholder
            "caseload_op": None,   # Placeholder
            "iic_capacity": iic_cap,
            "op_capacity": op_cap,
            "utilization_pct": util_pct,
        }

        if leader_id not in groups_map:
            groups_map[leader_id] = []
        groups_map[leader_id].append(therapist_row)

    # Build response groups
    # Fetch leader names
    leader_ids = [lid for lid in groups_map if lid]
    leader_names = {}
    leader_emp = {}
    if leader_ids:
        leaders = await sb_request("GET", "users", params={
            "id": f"in.({','.join(leader_ids)})",
            "select": "id,first_name,last_name,employment_status",
        }) or []
        for l in leaders:
            leader_names[l["id"]] = f"{l.get('first_name', '')} {l.get('last_name', '')}".strip()
            leader_emp[l["id"]] = l.get("employment_status", "full_time")

    groups = []
    for lid, therapists in groups_map.items():
        if lid is None:
            continue
        # Sum team totals
        team = {
            "iic": sum(t["iic"] for t in therapists),
            "op": sum(t["op"] for t in therapists),
            "sbys": sum(t["sbys"] for t in therapists),
            "ados": sum(t["ados"] for t in therapists),
            "sick": sum(t["sick"] for t in therapists),
            "pto": sum(t["pto"] for t in therapists),
            "total_hours": sum(t["total_hours"] for t in therapists),
            "avg_per_period": round(sum(t["total_hours"] for t in therapists) / num_periods, 1),
        }
        # % meeting threshold
        meeting = sum(1 for t in therapists if t["status"] == "on_track")
        pct = round(meeting / max(len(therapists), 1) * 100, 0)

        groups.append({
            "leader_id": lid,
            "leader_name": leader_names.get(lid, "Unknown"),
            "leader_employment_status": leader_emp.get(lid, "full_time"),
            "team_totals": team,
            "pct_meeting_threshold": pct,
            "therapists": sorted(therapists, key=lambda t: t["name"]),
        })

    # Add unassigned group at the end
    unassigned = groups_map.get(None, [])
    if unassigned:
        team = {
            "iic": sum(t["iic"] for t in unassigned),
            "op": sum(t["op"] for t in unassigned),
            "sbys": sum(t["sbys"] for t in unassigned),
            "ados": sum(t["ados"] for t in unassigned),
            "sick": sum(t["sick"] for t in unassigned),
            "pto": sum(t["pto"] for t in unassigned),
            "total_hours": sum(t["total_hours"] for t in unassigned),
            "avg_per_period": round(sum(t["total_hours"] for t in unassigned) / num_periods, 1),
        }
        meeting = sum(1 for t in unassigned if t["status"] == "on_track")
        pct = round(meeting / max(len(unassigned), 1) * 100, 0)
        groups.append({
            "leader_id": None,
            "leader_name": "Unassigned",
            "leader_employment_status": None,
            "team_totals": team,
            "pct_meeting_threshold": pct,
            "therapists": sorted(unassigned, key=lambda t: t["name"]),
        })

    # Sort groups: named leaders first (alphabetical), then unassigned
    groups.sort(key=lambda g: (g["leader_id"] is None, g["leader_name"]))

    return {
        "timeframe": timeframe,
        "period": period,
        "period_label": period_label,
        "available_periods": available_periods,
        "thresholds": {k: v["monthly"] for k, v in PERF_THRESHOLDS.items()},
        "groups": groups,
    }


@app.get("/api/analytics/performance/{user_id}")
async def analytics_performance_detail(user_id: str, user=Depends(verify_token)):
    """Performance detail for a single user: month-by-month breakdown."""
    profile = user
    caller_role = profile.get("role", "therapist")
    caller_id = profile.get("id")

    # Visibility check
    if caller_role not in ("admin",):
        if caller_role == "clinical_leader":
            # Can see self + supervisees
            target = await sb_request("GET", "users", params={"id": f"eq.{user_id}", "select": "id,clinical_supervisor_id"})
            if target and target[0].get("clinical_supervisor_id") != caller_id and target[0]["id"] != caller_id:
                raise HTTPException(status_code=403, detail="Not authorized")
        elif caller_id != user_id:
            raise HTTPException(status_code=403, detail="Not authorized")

    # Get user info
    target_user = await sb_request("GET", "users", params={
        "id": f"eq.{user_id}",
        "select": "id,first_name,last_name,role,employment_status",
    })
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    tu = target_user[0]
    emp_status = tu.get("employment_status", "full_time")
    threshold = PERF_THRESHOLDS.get(emp_status, PERF_THRESHOLDS["full_time"])["monthly"]

    # Get all closed pay periods, ordered by start_date
    all_periods = await sb_request("GET", "pay_periods", params={
        "select": "id,start_date,end_date,label,status",
        "status": "eq.closed",
        "order": "start_date.asc",
    }) or []

    # Group pay periods by month
    months_data = OrderedDict()
    for pp in all_periods:
        mk = pp["start_date"][:7]  # YYYY-MM
        if mk not in months_data:
            months_data[mk] = {"period_ids": [], "month_key": mk}
        months_data[mk]["period_ids"].append(pp["id"])

    # For each month, get time_entries for this user
    monthly_rows = []
    months_names = ['January','February','March','April','May','June','July','August','September','October','November','December']

    for mk, minfo in months_data.items():
        hrs = {"iic": 0, "op": 0, "sbys": 0, "ados": 0, "sick": 0, "pto": 0, "admin": 0, "total": 0}
        for pid in minfo["period_ids"]:
            entries = await sb_request("GET", "time_entries", params={
                "pay_period_id": f"eq.{pid}",
                "user_id": f"eq.{user_id}",
                "select": "quantity,rate_types(name)",
            }) or []
            for e in entries:
                rt = e.get("rate_types") or {}
                rname = rt.get("name", "")
                qty = float(e.get("quantity", 0))
                col = RATE_TO_PERF_COL.get(rname, None)
                if col and col in hrs:
                    hrs[col] += qty
                hrs["total"] += qty

        y2, m2 = int(mk[:4]), int(mk[5:7])
        on_track = hrs["total"] >= threshold
        monthly_rows.append({
            "month": mk,
            "label": f"{months_names[m2-1]} {y2}",
            "iic": round(hrs["iic"], 1),
            "op": round(hrs["op"], 1),
            "sbys": round(hrs["sbys"], 1),
            "ados": round(hrs["ados"], 1),
            "sick": round(hrs["sick"], 1),
            "pto": round(hrs["pto"], 1),
            "admin_hours": round(hrs["admin"], 1),
            "total_hours": round(hrs["total"], 1),
            "on_track": on_track,
        })

    # Build quarterly summaries
    quarters = OrderedDict()
    for row in monthly_rows:
        y2, m2 = int(row["month"][:4]), int(row["month"][5:7])
        qk = f"{y2}-Q{(m2-1)//3+1}"
        if qk not in quarters:
            quarters[qk] = {"quarter": qk, "months": [], "total_hours": 0}
        quarters[qk]["months"].append(row)
        quarters[qk]["total_hours"] += row["total_hours"]
    for qk, qdata in quarters.items():
        qdata["total_hours"] = round(qdata["total_hours"], 1)
        qdata["avg_per_month"] = round(qdata["total_hours"] / max(len(qdata["months"]), 1), 1)
        qdata["on_track"] = qdata["avg_per_month"] >= threshold

    return {
        "user": {
            "id": tu["id"],
            "name": f"{tu.get('first_name', '')} {tu.get('last_name', '')}".strip(),
            "role": tu.get("role", ""),
            "employment_status": emp_status,
        },
        "threshold": threshold,
        "months": list(reversed(monthly_rows)),  # Most recent first
        "quarters": list(reversed(list(quarters.values()))),
    }


@app.patch("/api/analytics/therapist-capacity/{user_id}")
async def update_therapist_capacity(user_id: str, req: dict = Body(...), admin=Depends(require_admin)):
    """Admin: set or update capacity targets for a therapist."""
    existing = await sb_request("GET", "therapist_capacity", params={
        "user_id": f"eq.{user_id}", "select": "id",
    })
    data = {
        "user_id": user_id,
        "iic_capacity": int(req.get("iic_capacity", 0) or 0),
        "op_capacity": int(req.get("op_capacity", 0) or 0),
        "updated_at": datetime.utcnow().isoformat(),
    }
    if existing:
        await sb_request("PATCH", f"therapist_capacity?user_id=eq.{user_id}", data=data)
    else:
        await sb_request("POST", "therapist_capacity", data=data)
    return {"status": "updated"}


# ── VTO (Vision/Traction Organizer) ──

@app.get("/api/vto")
async def get_vto(user=Depends(verify_token)):
    """Get VTO data stored in app_settings."""
    rows = await sb_request("GET", "app_settings", params={
        "key": "eq.vto_data",
        "select": "value",
    })
    if rows and rows[0].get("value"):
        import json as _json
        try:
            return _json.loads(rows[0]["value"])
        except Exception:
            return rows[0]["value"]
    # Return default VTO structure
    return {
        "org_name": "BestLife Counseling Services",
        "vision": {
            "core_values": [
                "Teamwork makes the dream work",
                "Support providers in any stage of their career",
                "Remove barriers and ensure continuity of care",
                "Identify an issue, propose a solution",
            ],
            "core_focus": {"passion": "Change the Game", "niche": "Empowering providers every step of the way"},
            "ten_year_target": "Impact 10,000 lives",
            "marketing_strategy": {
                "target_market": [
                    "Create a relationship with community agencies and supports that will give us referrals",
                    "Direct to consumer via social media",
                ],
                "three_uniques": [
                    "Technology forward solutions",
                    "Quick and accurate communication",
                    "Well connected in the community",
                ],
                "proven_process": "The BestLife Difference",
            },
            "three_year_picture": {
                "future_date": "December 31, 2027",
                "revenue": "$1.2 million",
                "profit": "$120,000",
                "measurables": "10% net margin",
                "what_does_it_look_like": [
                    "Best culture ever",
                    "Expanded benefits (401K, Holiday Pay)",
                    "Expanded benefits (Clinical Leaders)",
                    "Group therapy initiatives",
                    "10 full time clinicians",
                    "Cumberland County office",
                    "BestLife virtual hub",
                    "Fractional admin (marketing, finance, HR)",
                    "Expand APN services (2 new providers)",
                    "Med management",
                    "Intern/BA \u2192 Supervisor program",
                    "Better BestLife office",
                ],
            },
        },
        "traction": {
            "one_year_plan": {
                "future_date": "12/31/26",
                "revenue": "$1,029,000",
                "profit": "$26,600",
                "measurables": "2.59% net margin",
                "goals": [
                    "Run therapy groups",
                    "BestLife Virtual Hub",
                    "45% room utilization rate",
                    "Clinical Leader flow dialed in",
                    "Intake process streamlined",
                    "Website upgrade",
                    "RCM on a cadence",
                ],
            },
            "rocks": [
                {"title": "Initial group research complete", "who": "Adge"},
                {"title": "Intake process testing", "who": "Dave"},
                {"title": "Portal upgrades", "who": "Tim"},
            ],
            "issues": [],
        },
    }


@app.patch("/api/vto")
async def update_vto(req: dict = Body(...), admin=Depends(require_admin)):
    """Admin: Update VTO data in app_settings."""
    import json as _json
    value_str = _json.dumps(req)
    existing = await sb_request("GET", "app_settings", params={
        "key": "eq.vto_data", "select": "id",
    })
    if existing:
        await sb_request("PATCH", "app_settings?key=eq.vto_data", data={"value": value_str})
    else:
        await sb_request("POST", "app_settings", data={"key": "vto_data", "value": value_str})
    return {"status": "updated"}


# ── Impact Hours ──

@app.get("/api/impact-hours")
async def get_impact_hours(user=Depends(verify_token)):
    """Get total hours of impact (baseline + all time_entries)."""
    # Get baseline from app_settings
    rows = await sb_request("GET", "app_settings", params={
        "key": "eq.impact_hours_baseline",
        "select": "value",
    })
    baseline = 0
    if rows and rows[0].get("value"):
        try:
            baseline = float(rows[0]["value"])
        except Exception:
            pass

    # Sum all approved time_entries quantity
    # Use rollup_monthly for efficiency
    rollups = await sb_request("GET", "rollup_monthly", params={
        "select": "total_hours",
    }) or []
    total_from_entries = sum(float(r.get("total_hours", 0)) for r in rollups)

    return {"baseline": baseline, "from_entries": round(total_from_entries, 1), "total": round(baseline + total_from_entries, 1)}


@app.patch("/api/impact-hours")
async def update_impact_hours_baseline(req: dict = Body(...), admin=Depends(require_admin)):
    """Admin: Set the baseline impact hours."""
    baseline = float(req.get("baseline", 0))
    existing = await sb_request("GET", "app_settings", params={
        "key": "eq.impact_hours_baseline", "select": "id",
    })
    if existing:
        await sb_request("PATCH", "app_settings?key=eq.impact_hours_baseline", data={"value": str(baseline)})
    else:
        await sb_request("POST", "app_settings", data={"key": "impact_hours_baseline", "value": str(baseline)})
    return {"status": "updated", "baseline": baseline}


@app.get("/api/analytics/supervision-compliance")
async def analytics_supervision(user=Depends(verify_token)):
    """Analytics: Supervision compliance for clinical leaders."""
    profile = user
    role = profile.get("role")

    # Get users who require supervision
    params = {
        "supervision_required": "eq.true",
        "select": "id, first_name, last_name, clinical_supervisor_id",
    }

    # Scope: admin sees all, clinical_leader sees own supervisees, others see self
    if role == "clinical_leader":
        params["clinical_supervisor_id"] = f"eq.{profile['id']}"
    elif role != "admin":
        params["id"] = f"eq.{profile['id']}"

    supervisees = await sb_request("GET", "users", params=params)

    result = []
    for s in (supervisees or []):
        # Get supervisor name
        sup_name = "—"
        if s.get("clinical_supervisor_id"):
            sups = await sb_request("GET", "users", params={
                "id": f"eq.{s['clinical_supervisor_id']}",
                "select": "first_name,last_name",
            })
            if sups:
                sup_name = f"{sups[0]['first_name']} {sups[0]['last_name']}"

        result.append({
            "supervisee_name": f"{s['first_name']} {s['last_name']}",
            "supervisor_name": sup_name,
            "sessions_required": None,
            "sessions_completed": 0,
            "compliant": s.get("clinical_supervisor_id") is not None,
        })

    return result


# ── AI Features (Claude API) ─────────────────────────────────────

@app.get("/api/ai/status")
async def ai_status(admin=Depends(require_admin)):
    """Check if AI is configured (admin only)."""
    return {
        "configured": bool(ANTHROPIC_API_KEY),
        "key_prefix": ANTHROPIC_API_KEY[:12] + "..." if ANTHROPIC_API_KEY else None,
        "key_length": len(ANTHROPIC_API_KEY) if ANTHROPIC_API_KEY else 0,
        "source": "supabase" if not os.environ.get("ANTHROPIC_API_KEY") else "env",
    }


@app.post("/api/ai/chat")
async def ai_chat(req: AIChatRequest, user=Depends(verify_token)):
    """Send a prompt to Claude (Sonnet) and return the response."""
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=503, detail="AI is not configured. Add ANTHROPIC_API_KEY to environment or Supabase app_settings.")

    import anthropic

    system_message = req.system_hint or (
        "You are Betty, a helpful AI assistant for BestLife Behavioral Health. "
        "You help staff with questions about their dashboard, tasks, policies, billing, "
        "clinical workflows, and general practice operations. Be concise, friendly, and professional. "
        "If you don't know something specific to BestLife, say so honestly."
    )

    if req.context:
        system_message += f"\n\nAdditional context: {req.context}"

    try:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=req.max_tokens,
            system=system_message,
            messages=[{"role": "user", "content": req.prompt}],
        )
        # Extract text from response
        text = ""
        for block in message.content:
            if hasattr(block, "text"):
                text += block.text
        return {"response": text, "model": message.model, "usage": {"input_tokens": message.usage.input_tokens, "output_tokens": message.usage.output_tokens}}
    except anthropic.APIError as e:
        logger.error(f"Claude API error: {e}")
        raise HTTPException(status_code=502, detail=f"AI service error: {str(e)}")
    except Exception as e:
        logger.error(f"AI chat error: {e}")
        raise HTTPException(status_code=500, detail="Failed to get AI response")


@app.post("/api/ai/kb-assist")
async def ai_kb_assist(req: AIChatRequest, user=Depends(verify_token)):
    """AI-assisted content generation for Knowledge Base articles."""
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=503, detail="AI is not configured. Add ANTHROPIC_API_KEY to environment or Supabase app_settings.")

    import anthropic

    system_message = (
        "You are a content writer for BestLife Behavioral Health's internal knowledge base. "
        "Write clear, professional articles for therapists and staff. "
        "Use markdown formatting. Be thorough but concise. "
        "Structure content with headers, bullet points, and numbered lists where appropriate."
    )

    if req.context:
        system_message += f"\n\nArticle context — Category: {req.context}"

    try:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=req.max_tokens or 2048,
            system=system_message,
            messages=[{"role": "user", "content": req.prompt}],
        )
        text = ""
        for block in message.content:
            if hasattr(block, "text"):
                text += block.text
        return {"response": text, "model": message.model}
    except anthropic.APIError as e:
        logger.error(f"Claude API error: {e}")
        raise HTTPException(status_code=502, detail=f"AI service error: {str(e)}")
    except Exception as e:
        logger.error(f"AI KB assist error: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate content")


# ── Static File Serving (Production) ────────────────────────────

# Mount the built frontend in production
frontend_dist = os.path.join(os.path.dirname(__file__), "..", "frontend", "dist")
if os.path.exists(frontend_dist):
    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        file_path = os.path.join(frontend_dist, full_path)
        if os.path.isfile(file_path):
            return FileResponse(file_path)
        return FileResponse(os.path.join(frontend_dist, "index.html"))
