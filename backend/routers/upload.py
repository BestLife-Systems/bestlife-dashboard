"""TherapyNotes upload — 1 endpoint."""
import io
from datetime import datetime

import openpyxl
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends

from backend.deps import sb_request, require_admin, logger

router = APIRouter(prefix="/api")


@router.post("/upload/therapynotes")
async def upload_therapynotes(file: UploadFile = File(...), admin=Depends(require_admin)):
    """Process a TherapyNotes billing export and store transactions in Supabase."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active

    # Parse headers
    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {h: i for i, h in enumerate(headers)}

    required_cols = ["Record Type"]
    for col in required_cols:
        if col not in header_map:
            raise HTTPException(status_code=400, detail=f"Missing required column: {col}")

    # Parse rows
    transactions = []
    therapist_names = set()
    date_min = None
    date_max = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        def cell(name):
            idx = header_map.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        record_type = str(cell("Record Type") or "").strip()
        patient_name = str(cell("Patient Name") or cell("Patient") or "").strip()
        provider_name = str(cell("Provider Name") or cell("Provider") or "").strip()
        service_date = cell("Service Date") or cell("Date")
        amount = cell("Amount") or cell("Payment Amount") or 0
        payer = str(cell("Payer") or cell("Payer Name") or "").strip()
        code = str(cell("Code") or cell("Procedure Code") or "").strip()
        description = str(cell("Description") or "").strip()

        # Parse date
        date_str = None
        if isinstance(service_date, datetime):
            date_str = service_date.strftime("%Y-%m-%d")
        elif service_date:
            try:
                for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
                    try:
                        date_str = datetime.strptime(str(service_date).strip(), fmt).strftime("%Y-%m-%d")
                        break
                    except ValueError:
                        continue
            except Exception as e:
                logger.debug(f"Date parse failed for '{service_date}': {e}")

        # Parse amount
        try:
            amount_val = float(str(amount).replace("$", "").replace(",", "").replace("(", "-").replace(")", ""))
        except (ValueError, TypeError):
            amount_val = 0.0

        if provider_name:
            therapist_names.add(provider_name)

        if date_str:
            d = datetime.strptime(date_str, "%Y-%m-%d")
            if date_min is None or d < date_min:
                date_min = d
            if date_max is None or d > date_max:
                date_max = d

        transactions.append({
            "record_type": record_type,
            "patient_name": patient_name,
            "provider_name": provider_name,
            "service_date": date_str,
            "amount": amount_val,
            "payer": payer,
            "code": code,
            "description": description,
        })

    wb.close()

    if not transactions:
        raise HTTPException(status_code=400, detail="No transactions found in file")

    # Batch insert into Supabase (chunks of 500)
    chunk_size = 500
    for i in range(0, len(transactions), chunk_size):
        chunk = transactions[i:i + chunk_size]
        await sb_request("POST", "transactions", data=chunk)

    # Ensure therapists exist in users table (upsert check)
    for name in therapist_names:
        parts = name.split(None, 1)
        first = parts[0] if parts else name
        last = parts[1] if len(parts) > 1 else ""

        existing = await sb_request("GET", "therapists", params={
            "name": f"eq.{name}",
            "select": "id",
        })
        if not existing:
            await sb_request("POST", "therapists", data={
                "name": name,
                "first_name": first,
                "last_name": last,
            })

    # Update upload metadata
    await sb_request("POST", "upload_metadata", data={
        "filename": file.filename,
        "uploaded_by": admin["id"],
        "transactions_count": len(transactions),
        "therapist_count": len(therapist_names),
        "date_range_start": date_min.strftime("%Y-%m-%d") if date_min else None,
        "date_range_end": date_max.strftime("%Y-%m-%d") if date_max else None,
    })

    date_range = ""
    if date_min and date_max:
        date_range = f"{date_min.strftime('%b %d, %Y')} — {date_max.strftime('%b %d, %Y')}"

    return {
        "status": "success",
        "transactions_count": len(transactions),
        "therapist_count": len(therapist_names),
        "date_range": date_range,
    }
